"""
华阳精机简历筛选系统 · Web 应用 v2
=================================
绑定 0.0.0.0，支持局域网多人访问（一台机器抓取，多台机器查看/标注）。
"""

import os
import sys
import io
import json
import glob
import html
import shutil
import base64
import hashlib
import logging
import threading
import time
import random
import collections
import subprocess
import concurrent.futures
import re
import zipfile
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from flask import (Flask, render_template, request, jsonify,
                   redirect, url_for, send_file, Response)
from flask_cors import CORS

import database
import pipeline
import extractor
import worker
from utils import get_lan_ip
import jd_parser
import judger
import learning
import llm_utils
import init_jobs
import ranker
from browser_utils import _find_edge_exe, _write_profile_prefs, _snapshot_dir, _wait_for_new_file

try:
    import file_parser
    HAS_FILE_PARSER = True
except ImportError:
    HAS_FILE_PARSER = False


if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
for noisy in ["httpx", "httpcore", "openai", "urllib3", "selenium",
              "werkzeug", "chromadb"]:
    logging.getLogger(noisy).setLevel(logging.WARNING)
logger = logging.getLogger(__name__)


app = Flask(__name__)
CORS(app)

ROOT = os.path.dirname(os.path.abspath(__file__))

# 简历 PDF/Word 文件存储目录（按 resume_id 命名）
FILES_DIR = os.path.join(ROOT, "data", "resume_files")
os.makedirs(FILES_DIR, exist_ok=True)

# 人类可读命名的下载副本（M.DD-岗位-姓名.pdf），供服务器下载后直接浏览
NAMED_DL_DIR = os.path.join(ROOT, "data", "named_downloads")
os.makedirs(NAMED_DL_DIR, exist_ok=True)

# 服务端下载时的临时目录（Selenium 自动下载到此，然后移入 FILES_DIR）
DL_TMP_DIR = os.path.join(ROOT, "data", "dl_tmp")
os.makedirs(DL_TMP_DIR, exist_ok=True)

# 服务器本机 Edge 专用 Profile 目录（保存登录态 Cookie，避免每次重新登录）
EDGE_PROFILE_DIR = os.path.join(ROOT, "data", "edge_profile")
os.makedirs(EDGE_PROFILE_DIR, exist_ok=True)
EDGE_DEBUG_PORT = 19222


# =============================================================================
# 浏览器驱动管理（服务器本机，向后兼容）
# =============================================================================

_driver_state = {"driver": None, "job_name": None, "process": None, "attaching": False}
_driver_lock = threading.Lock()

# ── 预热：应用启动时后台解析 EdgeDriver，缓存路径供后续使用 ──────────────────
_prewarm_driver_future = None


def _prewarm_driver():
    global _prewarm_driver_future
    _ex = concurrent.futures.ThreadPoolExecutor(max_workers=1, thread_name_prefix="prewarm-srv")

    def _resolve():
        try:
            from webdriver_manager.microsoft import EdgeChromiumDriverManager
            path = EdgeChromiumDriverManager().install()
            logger.info(f"[预热] EdgeDriver 路径已缓存: {path}")
            return path
        except Exception as e:
            logger.info(f"[预热] EdgeDriver 预热失败（将使用 Selenium 内置管理）: {e}")
            return None

    _prewarm_driver_future = _ex.submit(_resolve)
    _ex.shutdown(wait=False)


_prewarm_driver()


def _browser_is_open():
    """浏览器是否处于开启状态（WebDriver 已连接 或 subprocess 进程仍在运行）"""
    with _driver_lock:
        drv  = _driver_state["driver"]
        proc = _driver_state.get("process")
    # .title 调用必须在锁外，避免 WebDriver 慢响应时长期持锁
    if drv:
        try:
            _ = drv.title
            return True
        except Exception:
            with _driver_lock:
                if _driver_state["driver"] is drv:
                    _driver_state["driver"] = None
    if proc and proc.poll() is None:
        return True
    return False

# 每次抓取完成后记录本次批次的 resume_id 集合，供"评估本批次"使用
_scrape_sessions: dict = {}

# =============================================================================
# 后台批量下载全局状态
# =============================================================================

_dl_status: dict = {
    "state":        "idle",   # idle | running | cancelled | done | error
    "job_name":     "",
    "filetype":     "pdf",
    "total":        0,
    "done":         0,
    "failed":       0,        # 51job 访问/CDP 失败的数量
    "current_name": "",
    "results":      [],       # [{resume_id, name, pdf_ok, fetch_err}]
    "error":        "",
    "started_at":   0.0,
    "completed_at": 0.0,
}
_dl_lock        = threading.Lock()
_dl_cancel_flag = threading.Event()

# 扩展设备本机批量下载任务状态（每台设备独立，互不影响）
_ext_dl_tasks: dict = {}  # device_id → {state,done,failed,total,current_name,job_name,batch_id,error_msg,started_at}
_ext_dl_lock  = threading.Lock()

_RESUME_ID_RE = re.compile(r'^[a-zA-Z0-9_\-]{4,80}$')

# =============================================================================
# 分布式抓取代理管理
# =============================================================================

_devices: dict = {}         # device_id → {name, last_seen, browser_open, status, current_job}
_device_commands: dict = {} # device_id → collections.deque of pending commands
_device_events: dict = {}   # device_id → threading.Event（长轮询等待唤醒）
_session_meta: dict = {}    # session_id → {device_name, job_name}
_agent_lock = threading.Lock()
DEVICE_TIMEOUT = 90         # 超过此秒数未心跳视为离线（需大于长轮询超时 25s）

# 服务器本机也作为一个内置设备（向后兼容）
SERVER_DEVICE_ID = "server-local"
SERVER_DEVICE_NAME = "服务器（本机）"

# =============================================================================
# 抓取任务队列（任何局域网用户均可提交，服务器串行执行）
# =============================================================================

_scrape_queue: collections.deque = collections.deque()
_queue_lock = threading.Lock()
_queue_worker_started = False


def _get_queue_snapshot():
    """返回当前队列的浅拷贝，供 API 返回"""
    with _queue_lock:
        return list(_scrape_queue)


def _queue_worker_loop():
    """后台长驻线程：有任务就取出执行，没有就短暂休眠"""
    while True:
        task = None
        with _queue_lock:
            if _scrape_queue:
                task = _scrape_queue.popleft()
        if task:
            jn = task["job_name"]
            pipeline._log(jn,
                f"▶ 队列任务开始（来自 {task['submitted_by']}，"
                f"目标 {task['target_count']} 份）")
            _run_full_flow(jn, task["target_count"],
                           task["auto_evaluate"],
                           close_browser_after=False)
        else:
            time.sleep(3)


def _ensure_queue_worker():
    global _queue_worker_started
    if _queue_worker_started:
        return
    _queue_worker_started = True
    t = threading.Thread(target=_queue_worker_loop,
                         daemon=True, name="scrape-queue-worker")
    t.start()


def _get_driver():
    with _driver_lock:
        return _driver_state["driver"]


def _set_driver(driver, job_name):
    with _driver_lock:
        _driver_state["driver"] = driver
        _driver_state["job_name"] = job_name


def _quit_driver():
    with _driver_lock:
        d = _driver_state["driver"]
        p = _driver_state.get("process")
        _driver_state["driver"] = None
        _driver_state["job_name"] = None
        _driver_state["process"] = None
        _driver_state["attaching"] = False
    if d:
        try:
            d.quit()
        except Exception:
            pass
    if p:
        try:
            # taskkill /F /T 同时终止子进程树（GPU/renderer 等 Chromium 子进程）
            subprocess.run(
                ["taskkill", "/F", "/T", "/PID", str(p.pid)],
                capture_output=True, timeout=5,
            )
        except Exception:
            try:
                p.terminate()
            except Exception:
                pass


def _attach_driver_background(port, job_name):
    """后台线程：等待 Edge CDP 端口就绪后附着 WebDriver（用户登录期间静默完成）"""
    import urllib.request as _urllib_req

    # 等待 Edge 调试端口就绪，最多 10 秒
    for _ in range(20):
        try:
            _urllib_req.urlopen(f"http://127.0.0.1:{port}/json/version", timeout=1)
            break
        except Exception:
            time.sleep(0.5)
    else:
        logger.warning(f"[附着] CDP 端口 {port} 等待超时，WebDriver 绑定失败")
        with _driver_lock:
            _driver_state["attaching"] = False
        return

    try:
        from selenium import webdriver
        from selenium.webdriver.edge.options import Options

        opts = Options()
        opts.add_experimental_option("debuggerAddress", f"127.0.0.1:{port}")

        drv = None
        if _prewarm_driver_future:
            try:
                driver_path = _prewarm_driver_future.result(timeout=60)
                if driver_path:
                    from selenium.webdriver.edge.service import Service
                    drv = webdriver.Edge(service=Service(driver_path), options=opts)
            except Exception as e:
                logger.debug(f"[附着] 使用预热路径失败: {e}")

        if drv is None:
            drv = webdriver.Edge(options=opts)

        # attach 模式下 prefs 无效；通过 CDP 补设下载行为
        try:
            drv.execute_cdp_cmd("Browser.setDownloadBehavior", {
                "behavior": "allow",
                "downloadPath": DL_TMP_DIR,
                "eventsEnabled": True,
            })
        except Exception:
            pass

        _set_driver(drv, job_name)
        logger.info(f"✓ [附着] WebDriver 已绑定到 Edge（端口 {port}）")
    except Exception as e:
        logger.error(f"[附着] WebDriver 绑定失败: {e}")
    finally:
        with _driver_lock:
            _driver_state["attaching"] = False


def _open_browser_server(job_name):
    """服务器本机：subprocess 瞬时启动 Edge，后台绑定 WebDriver"""
    with _driver_lock:
        existing_driver = _driver_state["driver"]
        existing_proc   = _driver_state.get("process")
        attaching       = _driver_state.get("attaching", False)

    # 驱动已连接且活跃
    if existing_driver:
        try:
            _ = existing_driver.title
            return jsonify(ok=True, message="浏览器已在运行")
        except Exception:
            pass  # driver 已死，继续重开

    # 进程在跑但 driver 未连（附着中 or 用户刚打开）
    if existing_proc and existing_proc.poll() is None:
        msg = "浏览器正在后台绑定驱动，请稍候..." if attaching else "浏览器已在运行"
        return jsonify(ok=True, message=msg)

    # 清理残留状态
    _quit_driver()

    edge_exe = _find_edge_exe()
    if not edge_exe:
        # 找不到 Edge 路径，回退到 Selenium 直接启动（后台线程）
        threading.Thread(target=_open_browser_fallback_server, args=(job_name,), daemon=True).start()
        return jsonify(ok=True, message="未找到 Edge 路径，正在后台启动浏览器，请稍候...")

    _write_profile_prefs(EDGE_PROFILE_DIR, DL_TMP_DIR)

    proc = subprocess.Popen([
        edge_exe,
        f"--remote-debugging-port={EDGE_DEBUG_PORT}",
        f"--user-data-dir={EDGE_PROFILE_DIR}",
        "--no-first-run",
        "--no-default-browser-check",
        "--start-maximized",
        "https://ehire.51job.com",
    ])

    with _driver_lock:
        _driver_state["process"]   = proc
        _driver_state["job_name"]  = job_name
        _driver_state["attaching"] = True

    threading.Thread(
        target=_attach_driver_background,
        args=(EDGE_DEBUG_PORT, job_name),
        daemon=True,
    ).start()

    return jsonify(ok=True, message="浏览器已打开，请登录 51job 并搜索本岗位关键词，然后点击「开始抓取并评估」")


def _open_browser_fallback_server(job_name):
    """Edge 可执行文件找不到时的后备：Selenium 直接启动（阻塞线程）"""
    try:
        from selenium import webdriver
        from selenium.webdriver.edge.options import Options
        opts = Options()
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("prefs", {
            "download.default_directory":        DL_TMP_DIR,
            "download.prompt_for_download":       False,
            "download.directory_upgrade":         True,
            "safebrowsing.enabled":               False,
            "plugins.always_open_pdf_externally": True,
        })
        driver_path = None
        if _prewarm_driver_future:
            try:
                driver_path = _prewarm_driver_future.result(timeout=60)
            except Exception:
                pass
        if driver_path:
            from selenium.webdriver.edge.service import Service
            drv = webdriver.Edge(service=Service(driver_path), options=opts)
        else:
            drv = webdriver.Edge(options=opts)
        drv.maximize_window()
        drv.get("https://ehire.51job.com")
        _set_driver(drv, job_name)
        logger.info("✓ [后备] 浏览器已启动（Selenium 直接启动模式）")
    except Exception as e:
        logger.error(f"[后备] 浏览器启动失败: {e}")


def _is_local_request():
    """抓取/浏览器控制仅允许服务器本机调用"""
    return request.remote_addr in ("127.0.0.1", "::1")


# =============================================================================
# 启动检查
# =============================================================================

def check_env():
    env_path = os.path.join(ROOT, ".env")
    if not os.path.exists(env_path):
        return False, ("未找到 .env 文件。请在程序目录下创建 .env，"
                       "内容：DEEPSEEK_API_KEY=你的密钥")
    from dotenv import load_dotenv
    load_dotenv(env_path)
    if not os.getenv("DEEPSEEK_API_KEY"):
        return False, (".env 中没有 DEEPSEEK_API_KEY。"
                       "请在 .env 文件中加上：DEEPSEEK_API_KEY=你的密钥")
    return True, ""


# =============================================================================
# 工具
# =============================================================================

def _safe_dir(name):
    return name.replace("/", "-").replace("\\", "-")


def _parse_structured(raw):
    """轻量解析 structured_json，返回 work_history 等前端导览卡需要的字段。"""
    if not raw:
        return None
    try:
        d = json.loads(raw) if isinstance(raw, str) else raw
        return {
            "work_history": d.get("work_history") or [],
            "industry_tags": d.get("industry_tags") or [],
        }
    except Exception:
        return None


def _parse_eval_item(e: dict):
    """统一解析 evaluation 记录的 matches/mismatches，消除多处重复代码。"""
    try:
        matches = json.loads(e["matches_json"]) if e.get("matches_json") else []
    except Exception:
        matches = []
    try:
        mismatches = json.loads(e["mismatches_json"]) if e.get("mismatches_json") else []
    except Exception:
        mismatches = []
    # 兼容旧格式 pros/cons
    if not matches and e.get("pros_json"):
        try:
            matches = [{"条件": "", "证据": p}
                       for p in json.loads(e["pros_json"]) if p]
        except Exception:
            pass
    if not mismatches and e.get("cons_json"):
        try:
            mismatches = [{"条件": "", "原因": c}
                          for c in json.loads(e["cons_json"]) if c]
        except Exception:
            pass
    return matches, mismatches


def _get_job_counts(job_name: str, job_id: int):
    """返回 (counts_dict, all_evals_full)，含预过滤排除数，供多处共用。"""
    all_evals = database.list_evaluations_for_job(job_name, include_hidden=True)
    pf = database.count_prefilter_rejects(job_id)
    counts = {
        "total":      len([e for e in all_evals if e["verdict"] != "排除"]),
        "deepgreen":  len([e for e in all_evals if e["verdict"] == "深绿"]),
        "lightgreen": len([e for e in all_evals if e["verdict"] == "蓝色"]),
        "yellow":     len([e for e in all_evals if e["verdict"] == "黄色"]),
        "excluded":   len([e for e in all_evals if e["verdict"] == "排除"]) + pf,
    }
    return counts, all_evals


def _resume_file_path(resume_id, filetype):
    """返回本地存储的 PDF 或 Word 文件路径（不保证文件存在）"""
    ext = ".pdf" if filetype == "pdf" else ".docx"
    return os.path.join(FILES_DIR, resume_id + ext)


def _build_named_filename(date_str, job_name, name, ext=".pdf"):
    """将 date_str(YYYY-MM-DD)、job_name、name 组合为 M.DD-岗位-姓名.pdf。
    任意字段缺失时返回 None，由调用方降级处理。"""
    def _safe(s, maxlen):
        return re.sub(r'[/\\:*?"<>|]', '_', str(s or '')).strip()[:maxlen]

    date_part = ''
    if date_str:
        parts = str(date_str).split('-')
        if len(parts) >= 3:
            try:
                m   = int(parts[1])
                day = int(parts[2])
                if m and day:
                    date_part = f"{m}.{day}"
            except (ValueError, IndexError):
                pass

    safe_job  = _safe(job_name, 20)
    safe_name = _safe(name, 20)
    if not (date_part and safe_job and safe_name):
        return None
    return f"{date_part}-{safe_job}-{safe_name}{ext}"


def _copy_named(date_str, job_name, name, src_path):
    """把 src_path 复制到 NAMED_DL_DIR，文件名用新格式；同名时追加 _2/_3 后缀。"""
    named = _build_named_filename(date_str, job_name, name)
    if not named:
        return
    dest    = os.path.join(NAMED_DL_DIR, named)
    counter = 2
    while os.path.exists(dest):
        stem = named[:-4]
        dest = os.path.join(NAMED_DL_DIR, f"{stem}_{counter}.pdf")
        counter += 1
    try:
        shutil.copy2(src_path, dest)
    except Exception:
        pass


def _check_resume_files(resume_ids):
    """批量检查哪些 resume_id 已有 PDF / Word 文件，返回两个 set"""
    has_pdf  = set()
    has_word = set()
    for rid in resume_ids:
        if os.path.exists(os.path.join(FILES_DIR, rid + ".pdf")):
            has_pdf.add(rid)
        if os.path.exists(os.path.join(FILES_DIR, rid + ".docx")):
            has_word.add(rid)
    return has_pdf, has_word


def _extract_browser_auth(drv):
    """
    从已登录浏览器提取完整鉴权状态：Cookie + localStorage + sessionStorage。
    现代 SPA（如51job ehire）将 Auth Token 存在 localStorage，仅转移 Cookie 不够。
    """
    cookies = drv.get_cookies()

    def _safe_storage(script):
        try:
            result = drv.execute_script(script)
            return result if isinstance(result, dict) else {}
        except Exception:
            return {}

    local_storage = _safe_storage(
        "var o={};"
        "try{ for(var i=0;i<localStorage.length;i++){"
        "  var k=localStorage.key(i); o[k]=localStorage.getItem(k);} }catch(e){}"
        "return o;")

    session_storage = _safe_storage(
        "var o={};"
        "try{ for(var i=0;i<sessionStorage.length;i++){"
        "  var k=sessionStorage.key(i); o[k]=sessionStorage.getItem(k);} }catch(e){}"
        "return o;")

    logger.info(f"[下载] 鉴权状态：Cookie {len(cookies)} 个，"
                f"localStorage {len(local_storage)} 项，"
                f"sessionStorage {len(session_storage)} 项")
    return cookies, local_storage, session_storage


def _create_headless_driver(cookies, local_storage, session_storage):
    """
    启动临时无头 Edge，克隆主浏览器的完整鉴权状态：
      1. Cookie（清除匿名 Session 后重新注入）
      2. localStorage（SPA Auth Token 的真正存储位置）
      3. sessionStorage
    注入后执行 refresh()，令 Vue.js SPA 重新读取 localStorage 恢复登录状态。
    """
    from selenium import webdriver
    from selenium.webdriver.edge.options import Options

    headless_profile = os.path.join(ROOT, "data", "headless_profile")
    os.makedirs(headless_profile, exist_ok=True)

    opts = Options()
    opts.add_argument("--headless")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument(f"--user-data-dir={headless_profile}")

    drv = None
    if _prewarm_driver_future:
        try:
            driver_path = _prewarm_driver_future.result(timeout=60)
            if driver_path:
                from selenium.webdriver.edge.service import Service
                drv = webdriver.Edge(service=Service(driver_path), options=opts)
        except Exception:
            pass
    if drv is None:
        drv = webdriver.Edge(options=opts)

    # ① 先导航到目标域名（localStorage 有域名作用域限制，必须先到该域）
    drv.get("https://ehire.51job.com")
    time.sleep(1.5)

    # ② 清除匿名 Session Cookie，注入已登录 Cookie
    drv.delete_all_cookies()
    for ck in cookies:
        try:
            c = {k: v for k, v in ck.items()
                 if k in ("name", "value", "domain", "path", "secure", "httpOnly")}
            drv.add_cookie(c)
        except Exception:
            pass

    # ③ 注入 localStorage（SPA Auth Token 的关键存储）
    if local_storage:
        drv.execute_script(
            "var d=arguments[0]; localStorage.clear();"
            "Object.entries(d).forEach(function(kv){"
            "  try{localStorage.setItem(kv[0],kv[1]);}catch(e){} });",
            local_storage)

    # ④ 注入 sessionStorage
    if session_storage:
        drv.execute_script(
            "var d=arguments[0]; sessionStorage.clear();"
            "Object.entries(d).forEach(function(kv){"
            "  try{sessionStorage.setItem(kv[0],kv[1]);}catch(e){} });",
            session_storage)

    # ⑤ 刷新页面，让 Vue.js SPA 重新初始化并读取 localStorage 中的 Token
    drv.refresh()
    time.sleep(2.5)

    return drv


def _bg_download_worker(items, filetype, job_name):
    """
    后台顺序下载简历（文字版 PDF）。
    原理：克隆主浏览器完整鉴权状态（Cookie+localStorage+sessionStorage）
         → 注入无头 Edge → Page.printToPDF（文字版 PDF）。
    """
    def _upd(**kw):
        with _dl_lock:
            _dl_status.update(kw)

    _upd(state="running", job_name=job_name, filetype="pdf",
         total=len(items), done=0, failed=0, current_name="",
         results=[], error="", started_at=time.time(), completed_at=0.0)

    results = []
    headless_drv = None

    try:
        main_drv = _get_driver()
        if not main_drv:
            _upd(state="error", error="浏览器已关闭，请先打开浏览器并登录 51job",
                 completed_at=time.time())
            return

        # 确保主浏览器在 ehire.51job.com 域（localStorage 域名需匹配）
        if "ehire.51job.com" not in main_drv.current_url:
            main_drv.get("https://ehire.51job.com")
            time.sleep(2.5)
        if "51job.com" not in main_drv.current_url:
            _upd(state="error", error="浏览器未登录 51job，请先登录",
                 completed_at=time.time())
            return

        cookies, local_storage, session_storage = _extract_browser_auth(main_drv)
        headless_drv = _create_headless_driver(cookies, local_storage, session_storage)

        for i, item in enumerate(items):
            if _dl_cancel_flag.is_set():
                _upd(state="cancelled", completed_at=time.time())
                return

            with _driver_lock:
                active_job = _driver_state.get("job_name")
            if active_job and pipeline.get_state(active_job).get("running"):
                _upd(state="error",
                     error=f"抓取「{active_job}」已启动，后台下载已中止以避免冲突",
                     completed_at=time.time())
                return

            resume_id = item["resume_id"]
            name      = item.get("name", resume_id)
            _upd(current_name=name)

            pdf_path   = os.path.join(FILES_DIR, resume_id + ".pdf")
            pdf_exists = os.path.exists(pdf_path)

            res = {"resume_id": resume_id, "name": name,
                   "pdf_ok": False, "fetch_err": False}

            if pdf_exists:
                res["pdf_ok"] = True
                results.append(res)
                _upd(done=i + 1, results=list(results))
                _copy_named(item.get("date", ""), job_name, name, pdf_path)
                continue

            try:
                from selenium.webdriver.common.by import By
                from selenium.webdriver.support.ui import WebDriverWait
                from selenium.webdriver.support import expected_conditions as EC

                url = (f"https://ehire.51job.com/Revision/talent/resume/detail"
                       f"?resumeId={resume_id}")
                headless_drv.get(url)
                # 等待简历正文区块出现（与抓取流程一致），页面渲染完即继续，
                # 避免固定 sleep(3) 的浪费；若 8 秒内未出现（极罕见），回退固定等待
                try:
                    WebDriverWait(headless_drv, 8).until(
                        EC.presence_of_element_located((By.ID, "work")))
                except Exception:
                    time.sleep(3)

                final_url = headless_drv.current_url
                if "resumeId" not in final_url or "detail" not in final_url:
                    raise RuntimeError(
                        f"Cookie 注入后仍被重定向至 {final_url}，"
                        "请关闭浏览器后重新打开并登录 51job"
                    )

                cdp_result = headless_drv.execute_cdp_cmd("Page.printToPDF", {
                    "printBackground": True,
                    "format": "A4",
                    "marginTop": 0.5, "marginBottom": 0.5,
                    "marginLeft": 0.4, "marginRight": 0.4,
                    "scale": 1.0,
                })
                pdf_bytes = base64.b64decode(cdp_result.get("data", ""))
                if not pdf_bytes or pdf_bytes[:4] != b"%PDF":
                    raise ValueError("打印结果不是有效 PDF")
                if len(pdf_bytes) > 50 * 1024 * 1024:
                    raise ValueError("PDF 超过 50MB，跳过")

                with open(pdf_path, "wb") as f:
                    f.write(pdf_bytes)
                res["pdf_ok"] = True
                _copy_named(item.get("date", ""), job_name, name, pdf_path)

            except Exception as e:
                logger.warning(f"后台下载失败 ({resume_id}): {e}")
                res["fetch_err"] = True

            results.append(res)
            failed = sum(1 for r in results if r.get("fetch_err"))
            _upd(done=i + 1, failed=failed, results=list(results))

            if i < len(items) - 1 and not _dl_cancel_flag.is_set():
                time.sleep(1.0)

    except Exception as e:
        logger.error(f"后台下载 worker 异常崩溃: {e}")
        _upd(state="error", error=f"下载异常终止：{e}",
             completed_at=time.time())
        return
    finally:
        if headless_drv:
            try:
                headless_drv.quit()
            except Exception:
                pass

    failed = sum(1 for r in results if r.get("fetch_err"))
    _upd(state="done", done=len(results), failed=failed,
         results=results, current_name="", completed_at=time.time())


def _gz_or_plain_path(raw_html_path):
    """根据 candidates.raw_html_path 推算实际 HTML 文件路径（支持 .gz）"""
    if not raw_html_path:
        return None
    if os.path.exists(raw_html_path):
        return raw_html_path
    if not raw_html_path.endswith(".gz") and os.path.exists(raw_html_path + ".gz"):
        return raw_html_path + ".gz"
    if raw_html_path.endswith(".gz"):
        plain = raw_html_path[:-3]
        if os.path.exists(plain):
            return plain
    return None


# =============================================================================
# 页面路由
# =============================================================================

@app.route("/")
def index():
    return redirect(url_for("dashboard"))


@app.route("/dashboard")
def dashboard():
    env_ok, env_msg = check_env()
    jobs = database.list_jobs()
    with database.get_conn() as conn:
        for j in jobs:
            r = conn.execute(
                "SELECT COUNT(*) AS c FROM evaluations WHERE job_id=?",
                (j["id"],)).fetchone()
            j["eval_count"] = r["c"] if r else 0
            apr = conn.execute("""
                SELECT COUNT(*) AS c FROM outcomes o
                JOIN evaluations e ON o.evaluation_id = e.id
                WHERE e.job_id=? AND o.action IN ('approved','hired')
                  AND NOT EXISTS (
                    SELECT 1 FROM outcomes o2
                    WHERE o2.evaluation_id = o.evaluation_id
                      AND o2.id > o.id
                      AND o2.action IN ('approved','hired','disapproved','rejected')
                  )
            """, (j["id"],)).fetchone()
            j["approved_count"] = apr["c"] if apr else 0
            dis = conn.execute("""
                SELECT COUNT(*) AS c FROM outcomes o
                JOIN evaluations e ON o.evaluation_id = e.id
                WHERE e.job_id=? AND o.action IN ('disapproved','rejected')
                  AND NOT EXISTS (
                    SELECT 1 FROM outcomes o2
                    WHERE o2.evaluation_id = o.evaluation_id
                      AND o2.id > o.id
                      AND o2.action IN ('approved','hired','disapproved','rejected')
                  )
            """, (j["id"],)).fetchone()
            j["disapproved_count"] = dis["c"] if dis else 0
    for j in jobs:
        ws = worker.get_job_status(j["name"])
        j["pending"]    = ws.get("pending", 0)
        j["html_total"] = ws.get("html_total", 0)
        j["is_running"] = ws.get("is_running", False)

    import datetime as _dt
    due_reminders = database.list_due_talent_pool(days_ahead=0)
    return render_template("dashboard.html",
                           jobs=jobs,
                           env_ok=env_ok, env_msg=env_msg,
                           due_reminders=due_reminders,
                           today_iso=_dt.date.today().isoformat())


@app.route("/cards/<path:job_name>")
def cards_page(job_name):
    return redirect(url_for("cockpit_page", job_name=job_name))


# =============================================================================
# 通过简历页面（需求7、8）
# =============================================================================

@app.route("/approved/<path:job_name>")
def approved_page(job_name):
    job = database.get_job(job_name)
    if not job:
        return f"岗位「{job_name}」不存在", 404
    highlight_raw = request.args.get("highlight", "")
    try:
        highlight_id = int(highlight_raw) if highlight_raw else None
    except (ValueError, TypeError):
        highlight_id = None
    # 带 highlight 时清空日期过滤，确保目标候选人一定在列表中
    if highlight_id:
        date_from = ""
        date_to   = ""
    else:
        date_from = request.args.get("date_from", "")
        date_to   = request.args.get("date_to", "")
    candidates = database.list_approved_for_job(
        job_name,
        date_from=date_from or None,
        date_to=date_to or None)

    # 注入文件标志
    rids = [c["resume_id"] for c in candidates]
    has_pdf, has_word = _check_resume_files(rids)
    for c in candidates:
        c["has_pdf"]  = c["resume_id"] in has_pdf
        c["has_word"] = c["resume_id"] in has_word

    # 注入 HR 阶段标签
    stage_map = database.get_hr_stages_for_job(job_name)
    for c in candidates:
        ht = stage_map.get(c["evaluation_id"]) or {}
        try:
            c["hr_stages"] = json.loads(ht.get("stages_json") or "[]")
        except Exception:
            c["hr_stages"] = []
        try:
            c["hr_stage_times"] = json.loads(ht.get("stage_times_json") or "{}")
        except Exception:
            c["hr_stage_times"] = {}
        c["department"] = ht.get("department") or ""
        c["hr_reject_reason"] = ht.get("reject_reason") or ""
        c["hr_note"] = ht.get("note") or ""
        # 解析 matches/mismatches
        matches, mismatches = _parse_eval_item(c)
        c["matches"]    = matches
        c["mismatches"] = mismatches

    # 阶段统计
    stage_counts = database.count_hr_stages_for_job(job_name)

    # 硬性条件名（用于前端过滤 matches/mismatches 中的冗余项）
    hard_cond_names = sorted(_get_hard_condition_names(job))

    return render_template(
        "approved.html",
        job_name=job_name,
        job_id=job["id"],
        candidates=candidates,
        stage_order=database.HR_STAGE_ORDER,
        reject_reasons=database.REJECT_REASONS,
        stage_counts=stage_counts,
        date_from=date_from,
        date_to=date_to,
        total=len(candidates),
        hard_cond_names=hard_cond_names,
        highlight_id=highlight_id,
    )


# =============================================================================
# 不通过简历页面（需求7、9）
# =============================================================================

@app.route("/disapproved/<path:job_name>")
def disapproved_page(job_name):
    job = database.get_job(job_name)
    if not job:
        return f"岗位「{job_name}」不存在", 404
    candidates = database.list_disapproved_for_job(job_name)

    rids = [c["resume_id"] for c in candidates]
    has_pdf, has_word = _check_resume_files(rids)
    stage_map = database.get_hr_stages_for_job(job_name)
    for c in candidates:
        c["has_pdf"]  = c["resume_id"] in has_pdf
        c["has_word"] = c["resume_id"] in has_word
        ht = stage_map.get(c["evaluation_id"]) or {}
        c["hr_note"] = ht.get("note") or ""
        matches, mismatches = _parse_eval_item(c)
        c["matches"]    = matches
        c["mismatches"] = mismatches

    return render_template(
        "disapproved.html",
        job_name=job_name,
        candidates=candidates,
        total=len(candidates),
    )


@app.route("/triage/<path:job_name>")
def triage_page(job_name):
    """快速处理台 - 主操作入口"""
    job = database.get_job(job_name)
    if not job:
        return f"岗位「{job_name}」不存在", 404

    job_config = json.loads(job["config_json"])

    # 默认权重
    weights = ranker.DEFAULT_WEIGHTS.copy()
    target_years = job_config.get("min_years", 5) or 5
    industries_required = []
    ind_raw = job_config.get("industry_required")
    if isinstance(ind_raw, list):
        industries_required = ind_raw
    elif isinstance(ind_raw, str) and ind_raw.strip():
        industries_required = [s.strip() for s in ind_raw.split("、") if s.strip()]

    # 取该岗位所有非排除候选人
    all_evals = database.list_evaluations_for_job(job_name, include_hidden=False)

    # 顶部统计
    counts = {
        "deepgreen": len([e for e in all_evals if e["verdict"] == "深绿"]),
        "lightgreen": len([e for e in all_evals if e["verdict"] == "蓝色"]),
        "yellow": len([e for e in all_evals if e["verdict"] == "黄色"]),
        "total": len(all_evals),
    }

    # 已处理 = 已有 approved 或 disapproved 的
    handled = sum(1 for e in all_evals
                  if e.get("latest_action") in ("approved", "disapproved",
                                                  "hired", "rejected"))

    # 历史淘汰标签（供下拉框）
    rejection_tag_options = database.get_distinct_rejection_tags(limit=30)

    return render_template(
        "triage.html",
        job_name=job_name,
        counts=counts,
        handled=handled,
        target_years=target_years,
        default_weights=weights,
        rejection_tag_options=rejection_tag_options,
        industries_required=industries_required,
    )


@app.route("/scrape/<path:job_name>")
def scrape_page(job_name):
    return redirect(url_for("cockpit_page", job_name=job_name))


@app.route("/cockpit/<path:job_name>")
def cockpit_page(job_name):
    """一体化驾驶舱：抓取控制 + 实时状态 + 候选人列表"""
    job = database.get_job(job_name)
    if not job:
        return f"岗位「{job_name}」不存在", 404
    config = json.loads(job["config_json"])
    default_count = config.get("target_count", 30)
    is_local = _is_local_request()
    return render_template("cockpit.html",
                           job_name=job_name,
                           default_count=default_count,
                           is_local=is_local)


@app.route("/prompt/<path:job_name>")
def prompt_page(job_name):
    job = database.get_job(job_name)
    if not job:
        return f"岗位「{job_name}」不存在", 404
    versions = _list_prompt_versions(job_name)
    return render_template("prompt_editor.html",
                           job_name=job_name,
                           prompt_text=job["prompt_text"] or "",
                           versions=versions)


@app.route("/database")
def database_page():
    tab = request.args.get("tab", "jobs")
    page_data = _query_database_table(tab,
                                      page=int(request.args.get("page", 1)),
                                      keyword=request.args.get("q", ""))
    return render_template("database.html", tab=tab, page_data=page_data)





@app.route("/cleanup/<path:job_name>")
def cleanup_page(job_name):
    """语义分组清理助手 - 仅对 verdict=排除 的候选人"""
    job = database.get_job(job_name)
    if not job:
        return f"岗位「{job_name}」不存在", 404
    return render_template("cleanup.html", job_name=job_name)


@app.route("/funnel")
def funnel_page():
    data = learning.get_funnel_page_data()
    return render_template("funnel.html", **data)


@app.route("/funnel/<path:job_name>")
def funnel_detail_page(job_name):
    detail = learning.stage4_funnel_detail(job_name)
    rarity = learning.stage5_rarity_for_job(job_name, sample_floor=10)
    reverse_jd = learning.stage6_reverse_jd_optimizer(job_name)
    supply_demand = learning.supply_demand_matrix(job_name)
    return render_template("funnel_detail.html",
                           job_name=job_name,
                           detail=detail,
                           rarity=rarity,
                           reverse_jd=reverse_jd,
                           supply_demand=supply_demand)


@app.route("/new-job")
def new_job_page():
    return render_template("new_job.html")


@app.route("/learning")
def learning_page():
    data = learning.get_learning_page_data()
    return render_template("learning.html", **data)


@app.route("/audit/<path:job_name>")
def audit_page(job_name):
    """筛选流水线诊断报告"""
    data = learning.get_audit_page_data(job_name)
    if not data:
        return f"岗位「{job_name}」不存在", 404
    return render_template("audit.html", **data)


@app.route("/criteria/<path:job_name>")
def criteria_page(job_name):
    """评估细则管理页面"""
    data = learning.get_criteria_page_data(job_name)
    if not data:
        return f"岗位「{job_name}」不存在", 404
    return render_template("criteria.html", **data)


@app.route("/api/correction/submit", methods=["POST"])
def api_correction_submit():
    """
    提交单条纠错信号。
    Body: {evaluation_id, direction, condition_name, error_type, evidence_text, hr_note}
    direction: 'too_loose' | 'too_strict'
    error_type: 'evidence_insufficient' | 'criterion_misunderstood' | 'criterion_not_important'
    """
    data = request.get_json() or {}
    eval_id = data.get("evaluation_id")
    direction = data.get("direction", "")
    condition_name = (data.get("condition_name") or "").strip()
    error_type = data.get("error_type", "")
    evidence_text = (data.get("evidence_text") or "").strip()[:200]
    hr_note = (data.get("hr_note") or "").strip()[:50]

    if not eval_id or direction not in ("too_loose", "too_strict") or not condition_name:
        return jsonify(ok=False, message="缺少必要参数"), 400

    try:
        with database.get_conn() as conn:
            row = conn.execute(
                "SELECT job_id FROM evaluations WHERE id=?", (int(eval_id),)
            ).fetchone()
        if not row:
            return jsonify(ok=False, message="评估记录不存在"), 404

        job_id = row["job_id"]
        database.record_correction_signal(
            job_id=job_id,
            eval_id=int(eval_id),
            direction=direction,
            condition_name=condition_name,
            error_type=error_type or None,
            evidence_text=evidence_text or None,
            hr_note=hr_note or None,
        )

        # 每积累 3 次未分析信号时，后台触发细则草稿生成
        counts = database.get_condition_correction_counts(job_id)
        if counts:
            job_row = database.get_job_by_id(job_id)
            job_name_bg = job_row["name"] if job_row else ""

            def _run_analysis(jid, jname):
                try:
                    client = llm_utils.initialize_client()
                    if client:
                        learning.analyze_correction_signals(jid, jname, client)
                except Exception as ex:
                    logger.warning(f"细则分析后台失败: {ex}")

            threading.Thread(
                target=_run_analysis,
                args=(job_id, job_name_bg),
                daemon=True,
                name="correction-analysis",
            ).start()

        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 500


@app.route("/api/criteria/list/<path:job_name>")
def api_criteria_list(job_name):
    """返回岗位评估细则列表（含 pending/confirmed/dismissed 分组）"""
    data = learning.get_criteria_page_data(job_name)
    if not data:
        return jsonify(ok=False, message="岗位不存在"), 404
    return jsonify(ok=True, **data)


@app.route("/api/criteria/confirm/<int:note_id>", methods=["POST"])
def api_criteria_confirm(note_id):
    """HR 确认一条细则草稿（可同时编辑文本和硬性标记）"""
    data = request.get_json() or {}
    note_text = (data.get("note_text") or "").strip()
    is_hard = int(bool(data.get("is_hard", False)))
    if not note_text:
        return jsonify(ok=False, message="细则内容不能为空"), 400
    try:
        database.confirm_criteria_note(note_id, note_text, is_hard)
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 500


@app.route("/api/criteria/dismiss/<int:note_id>", methods=["POST"])
def api_criteria_dismiss(note_id):
    """HR 驳回一条细则草稿"""
    try:
        database.dismiss_criteria_note(note_id)
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 500


@app.route("/api/criteria/refresh/<int:note_id>", methods=["POST"])
def api_criteria_refresh(note_id):
    """HR 复查后重置细则的 30 天计时"""
    try:
        database.refresh_criteria_note_reviewed(note_id)
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 500


@app.route("/api/audit/refresh/<path:job_name>", methods=["POST"])
def api_audit_refresh(job_name):
    """手动触发纠错信号分析，生成评估细则草稿"""
    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message="岗位不存在"), 404
    client = llm_utils.initialize_client()
    if not client:
        return jsonify(ok=False, message="LLM 不可用"), 500
    try:
        learning.analyze_correction_signals(job["id"], job_name, client)
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 500


# =============================================================================
# 内部辅助：获取该岗位的硬性条件名集合（用于前端过滤 matches/mismatches）
# =============================================================================

def _get_hard_condition_names(job):
    """从 job_criteria_notes(is_hard=1) 和 config_json.rules(type=hard) 提取硬性条件名。"""
    names = set()
    if not job:
        return names
    for n in database.get_confirmed_criteria_notes(job["id"]):
        if n.get("is_hard"):
            cname = (n.get("condition_name") or "").strip()
            if cname:
                names.add(cname)
    try:
        cfg = json.loads(job.get("config_json") or "{}")
        for r in cfg.get("rules", []):
            if r.get("type") == "hard":
                rname = (r.get("name") or "").strip()
                if rname:
                    names.add(rname)
    except Exception:
        pass
    return names


# =============================================================================
# API：候选人改名
# =============================================================================

@app.route("/api/candidate/rename", methods=["POST"])
def api_candidate_rename():
    data = request.get_json() or {}
    resume_id = (data.get("resume_id") or "").strip()
    new_name  = (data.get("new_name")  or "").strip()
    if not resume_id:
        return jsonify(ok=False, message="缺少 resume_id"), 400
    if not new_name:
        return jsonify(ok=False, message="姓名不能为空"), 400
    if len(new_name) > 50:
        return jsonify(ok=False, message="姓名不能超过50字"), 400
    try:
        with database.get_conn() as conn:
            result = conn.execute(
                "UPDATE candidates SET name=? WHERE resume_id=?",
                (new_name, resume_id))
            if result.rowcount == 0:
                return jsonify(ok=False, message="候选人不存在"), 404
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 500


# =============================================================================
# API：候选人操作记录（含通过/不通过）
# =============================================================================

@app.route("/api/outcome", methods=["POST"])
def api_outcome():
    data = request.get_json() or {}
    eval_id = data.get("evaluation_id")
    action = data.get("action")
    note = data.get("note")
    rejection_tag = data.get("rejection_tag")
    dwell_seconds = data.get("dwell_seconds")
    if not eval_id or not action:
        return jsonify(ok=False, message="缺少 evaluation_id 或 action"), 400
    try:
        database.record_outcome(int(eval_id), action, note, dwell_seconds)

        # 不通过时若带 rejection_tag，写入 rejection_tags 表
        with database.get_conn() as conn:
            eval_row = conn.execute(
                "SELECT job_id, verdict FROM evaluations WHERE id=?",
                (int(eval_id),)).fetchone()
        if eval_row:
            job_id = eval_row["job_id"]
            if rejection_tag and action in ("disapproved", "rejected"):
                database.add_rejection_tag(int(eval_id), job_id, rejection_tag)

        learning.invalidate_cache()
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 500


@app.route("/api/outcome/note", methods=["POST"])
def api_outcome_note():
    """更新评估的最新备注"""
    data = request.get_json() or {}
    eval_id = data.get("evaluation_id")
    note = data.get("note", "")
    if not eval_id:
        return jsonify(ok=False, message="缺少 evaluation_id"), 400
    try:
        database.update_latest_note(int(eval_id), note)
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 500


@app.route("/api/hr_stage", methods=["POST"])
def api_hr_stage():
    """保存候选人 HR 阶段标签（单选 + 时间存储 + 部门）"""
    data = request.get_json() or {}
    eval_id       = data.get("evaluation_id")
    job_id        = data.get("job_id")
    stage         = data.get("stage", "")          # 单个阶段字符串
    stage_time    = data.get("stage_time", "")     # 该阶段对应时间
    department    = data.get("department")          # None 表示不更新
    reject_reason = data.get("reject_reason", "")
    note          = data.get("note", "")

    if not eval_id or not job_id:
        return jsonify(ok=False, message="缺少 evaluation_id 或 job_id"), 400
    try:
        # 读取现有 stage_times_json
        existing = database.get_hr_stage(int(eval_id))
        if existing:
            try:
                times = json.loads(existing.get('stage_times_json') or '{}')
            except Exception:
                times = {}
        else:
            times = {}

        # 写入当前阶段时间
        if stage and stage_time:
            times[stage] = stage_time

        stages_list = [stage] if stage else []
        database.upsert_hr_stage(
            int(eval_id), int(job_id),
            stages=stages_list,
            reject_reason=reject_reason,
            note=note,
            stage_times_json=times,
            department=department,
        )
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 500


@app.route("/api/hr_stage/note", methods=["POST"])
def api_hr_stage_note():
    """仅更新不通过页备注（需求9），不影响 AI 分析"""
    data = request.get_json() or {}
    eval_id = data.get("evaluation_id")
    note = data.get("note", "")
    if not eval_id:
        return jsonify(ok=False, message="缺少 evaluation_id"), 400
    try:
        database.update_hr_note(int(eval_id), note)
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 500


@app.route("/api/hr_stage/counts/<path:job_name>")
def api_hr_stage_counts(job_name):
    """返回该岗位各 HR 阶段的实时人数（需求8(3)）"""
    counts = database.count_hr_stages_for_job(job_name)
    return jsonify(ok=True, counts=counts)


# =============================================================================
# API：处理台数据 + 跨岗位状态 + 简历 HTML 内容
# =============================================================================



@app.route("/api/triage/list", methods=["GET"])
def api_triage_list():
    """返回处理台需要的候选人列表（含权重排序）"""
    job_name = request.args.get("job_name", "")
    weights_str = request.args.get("weights")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name"), 400

    weights = ranker.DEFAULT_WEIGHTS.copy()
    if weights_str:
        try:
            wp = json.loads(weights_str)
            for k, v in wp.items():
                if k in weights:
                    weights[k] = float(v)
        except Exception:
            pass

    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message="岗位不存在"), 404
    job_config = json.loads(job["config_json"])
    target_years = job_config.get("min_years", 5) or 5
    industries_required = []
    ind_raw = job_config.get("industry_required")
    if isinstance(ind_raw, list):
        industries_required = ind_raw
    elif isinstance(ind_raw, str) and ind_raw.strip():
        industries_required = [s.strip() for s in ind_raw.split("、") if s.strip()]

    verdict_filter = request.args.get("verdict_filter", "")
    # 服务端推导 include_hidden：排除 Tab 强制为 True，无需依赖前端额外传参
    include_excluded_param = request.args.get("include_excluded", "0") == "1"
    include_hidden = include_excluded_param or (verdict_filter == "排除")
    exclude_processed = request.args.get("exclude_processed", "0") == "1"
    all_evals = database.list_evaluations_for_job(job_name,
                                                   include_hidden=include_hidden,
                                                   exclude_processed=exclude_processed)
    # 按 verdict_filter 过滤（匹配驾驶舱各 Tab 的显示逻辑）
    if verdict_filter == "全部":
        all_evals = [e for e in all_evals if e["verdict"] not in ("排除", "自投", "主动搜索")]
    elif verdict_filter in ("排除", "深绿", "蓝色", "黄色"):
        all_evals = [e for e in all_evals if e["verdict"] == verdict_filter]

    # 注入 structured_json 用于 ranker（一次批量查询，避免 N 次开关连接）
    if all_evals:
        cand_ids = [e["candidate_id"] for e in all_evals]
        ph = ",".join("?" * len(cand_ids))
        with database.get_conn() as conn:
            rows = conn.execute(
                f"SELECT id, structured_json FROM candidates WHERE id IN ({ph})",
                cand_ids).fetchall()
        sj_map = {r["id"]: r["structured_json"] for r in rows}
        for e in all_evals:
            e["structured_json"] = sj_map.get(e["candidate_id"])
    items = all_evals

    ranked = ranker.rank_candidates(
        items, weights,
        required_industries=industries_required,
        target_years=target_years,
    )

    out = []
    for e in ranked:
        matches, mismatches = _parse_eval_item(e)
        sj_raw = e.get("structured_json")
        out.append({
            "evaluation_id": e["id"],
            "candidate_id": e["candidate_id"],
            "resume_id": e["resume_id"],
            "name": e["name"],
            "age": e["age"],
            "first_degree": e["first_degree"],
            "school": e["school"],
            "major": e["major"],
            "english_level": e["english_level"],
            "total_years": e["total_years"],
            "verdict": e["verdict"],
            "verdict_reason": e.get("verdict_reason") or "",
            "matches": matches,
            "mismatches": mismatches,
            "latest_action": e.get("latest_action"),
            "latest_note": e.get("latest_note") or "",
            "rank_score": e.get("_rank_score", 0),
            "duplicate_of_id": e.get("duplicate_of_id"),
            "scrape_session_id": e.get("scrape_session_id"),
            "_structured": _parse_structured(sj_raw),
        })

    # 注入文件存在标志
    all_resume_ids = [item["resume_id"] for item in out]
    has_pdf, has_word = _check_resume_files(all_resume_ids)
    for item in out:
        item["has_pdf"]  = item["resume_id"] in has_pdf
        item["has_word"] = item["resume_id"] in has_word

    # 注入 HR 阶段标签和 decision_at（供前端按通过页顺序排序）
    eval_ids = [item["evaluation_id"] for item in out]
    stage_map = database.get_hr_stages_for_job(job_name)
    decision_at_map = {}
    if eval_ids:
        ph = ",".join("?" * len(eval_ids))
        with database.get_conn() as conn:
            rows = conn.execute(
                f"SELECT evaluation_id, MAX(action_at) AS decision_at "
                f"FROM outcomes WHERE evaluation_id IN ({ph}) "
                f"AND action IN ('approved','hired') "
                f"GROUP BY evaluation_id",
                eval_ids
            ).fetchall()
        decision_at_map = {r["evaluation_id"]: r["decision_at"] for r in rows}
    for item in out:
        eid = item["evaluation_id"]
        ht = stage_map.get(eid) or {}
        try:
            item["hr_stages"] = json.loads(ht.get("stages_json") or "[]")
        except Exception:
            item["hr_stages"] = []
        item["hr_reject_reason"] = ht.get("reject_reason") or ""
        item["decision_at"] = decision_at_map.get(eid)

    # 注入已确认评估细则
    job_id = job["id"] if job else None
    confirmed_notes = database.get_confirmed_criteria_notes(job_id) if job_id else []
    criteria_notes = [
        {
            "condition_name": n["condition_name"],
            "note_text": n["note_text"],
            "is_hard": bool(n["is_hard"]),
        }
        for n in confirmed_notes
    ]

    # 硬性条件名（供前端过滤 matches/mismatches 中的冗余项）
    hard_condition_names = sorted(_get_hard_condition_names(job))

    return jsonify(ok=True, candidates=out, count=len(out),
                   criteria_notes=criteria_notes,
                   hard_condition_names=hard_condition_names,
                   stage_order=database.HR_STAGE_ORDER)


@app.route("/api/triage/cross_jobs/<resume_id>", methods=["GET"])
def api_cross_jobs(resume_id):
    exclude = request.args.get("exclude_job_name")
    rows = database.get_candidate_cross_job_status(resume_id, exclude_job_name=exclude)
    return jsonify(ok=True, items=rows)


@app.route("/api/triage/transfer_job", methods=["POST"])
def api_triage_transfer_job():
    """将候选人从当前岗位转移到目标岗位并标为通过，原岗位自动标为不通过。
    首次调用若目标岗位已有评估记录返回 conflict=True，携带 force=true 强制覆盖。
    """
    data = request.get_json() or {}
    evaluation_id = data.get("evaluation_id")
    target_job_name = data.get("target_job_name", "").strip()
    verdict = data.get("verdict", "")
    force = bool(data.get("force", False))

    if not evaluation_id or not target_job_name or not verdict:
        return jsonify(ok=False, message="参数缺失"), 400
    if verdict not in ("深绿", "蓝色", "黄色"):
        return jsonify(ok=False, message="无效 verdict"), 400

    # 取源评估
    src_eval = database.get_evaluation_by_id(int(evaluation_id))
    if not src_eval:
        return jsonify(ok=False, message="源评估记录不存在"), 404
    candidate_id = src_eval["candidate_id"]

    # 取目标岗位
    target_job = database.get_job(target_job_name)
    if not target_job:
        return jsonify(ok=False, message=f"目标岗位「{target_job_name}」不存在"), 404
    target_job_id = target_job["id"]

    # 取源岗位名称
    src_job = database.get_job_by_id(src_eval["job_id"])
    src_job_name = src_job["name"] if src_job else "原岗位"

    # 检查目标岗位是否已有该候选人评估
    existing = database.existing_evaluation(candidate_id, target_job_id)
    if existing and not force:
        existing_action = database.get_latest_outcome_action(existing["id"])
        return jsonify(
            ok=False,
            conflict=True,
            existing_verdict=existing["verdict"],
            existing_action=existing_action or "",
        )

    # 执行转移
    if existing:
        # 已有评估：仅更新 verdict，保留 AI 分析
        database.update_evaluation_verdict(
            existing["id"], verdict,
            f"HR 手动转移（来自「{src_job_name}」）"
        )
        new_eval_id = existing["id"]
    else:
        # 新建评估
        new_eval_id = database.upsert_evaluation(
            candidate_id=candidate_id,
            job_id=target_job_id,
            verdict=verdict,
            pros=[],
            cons=[],
            has_hard_fail=False,
            matches=[],
            mismatches=[],
            verdict_reason=f"HR 手动转移（来自「{src_job_name}」）",
            session_id=None,
        )

    # 目标岗位标为 approved
    database.record_outcome(
        new_eval_id, "approved",
        note=f"由「{src_job_name}」转移"
    )

    # 源岗位标为 disapproved（不触发前端偏好弹窗）
    database.record_outcome(
        int(evaluation_id), "disapproved",
        note=f"已转移至「{target_job_name}」"
    )

    return jsonify(
        ok=True,
        redirect_url=f"/approved/{target_job_name}",
        message=f"已将候选人转移至「{target_job_name}」并标为通过"
    )


_RESUME_TRIAGE_STYLE = """<style id="_hy_triage_clean">
noscript { display: none !important; }
.navBar_wrap { display: none !important; }
.left_item_status_card { display: none !important; }
.resume_detail_right { display: none !important; }
.resume_detail_left { width: 100% !important; flex: 1 1 auto !important; max-width: 100% !important; }
.talent_resume_detail_wrap, .eh_body_min_width_resume { min-width: 0 !important; }
.ehire_gaea_html .eh_login, .ehire_gaea_html header main, .ehire_gaea_html .content { min-width: 0 !important; }
html, body { min-width: 0 !important; overflow-x: hidden !important; padding-top: 0 !important; }
</style>"""


@app.route("/api/resume_html/<resume_id>", methods=["GET"])
def api_resume_html(resume_id):
    """返回完整简历的 HTML 文本（用于处理台右栏内嵌渲染）"""
    cand = database.get_candidate_by_resume_id(resume_id)
    if not cand:
        return Response("候选人不存在", status=404)
    raw_path = cand.get("raw_html_path")
    actual = _gz_or_plain_path(raw_path) if raw_path else None
    if not actual:
        return Response("简历 HTML 文件不存在或已删除", status=404)
    html = pipeline.read_html_any(actual)
    if not html:
        return Response("读取失败", status=500)
    return Response(html + _RESUME_TRIAGE_STYLE, mimetype="text/html; charset=utf-8")


# =============================================================================
# API：候选人删除（处理台 D 键直接物理删除）
# =============================================================================

@app.route("/api/candidate/delete", methods=["POST"])
def api_candidate_delete():
    data = request.get_json() or {}
    candidate_id = data.get("candidate_id")
    resume_id = data.get("resume_id")
    if not candidate_id and not resume_id:
        return jsonify(ok=False, message="缺少 candidate_id 或 resume_id"), 400
    try:
        if not candidate_id and resume_id:
            cand = database.get_candidate_by_resume_id(resume_id)
            if not cand:
                return jsonify(ok=False, message="候选人不存在"), 404
            candidate_id = cand["id"]
        cand = database.get_candidate_by_id(int(candidate_id))
        n = database.delete_candidate(int(candidate_id))
        return jsonify(ok=True, deleted=n)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 500


# =============================================================================
# API：自然语言查询（NL2SQL，保留）
# =============================================================================


# =============================================================================
# API：岗位列表
# =============================================================================

@app.route("/api/jobs/list", methods=["GET"])
def api_jobs_list():
    jobs = database.list_jobs()
    return jsonify(ok=True, jobs=[{"id": j["id"], "name": j["name"]} for j in jobs])


# API：JD 创建岗位
# =============================================================================

@app.route("/api/jobs/create", methods=["POST"])
def api_create_job():
    data = request.get_json() or {}
    name = (data.get("name") or "").strip()
    jd = (data.get("jd") or "").strip()
    if not name or not jd:
        return jsonify(ok=False, message="岗位名或 JD 为空")
    client = llm_utils.initialize_client()
    if not client:
        return jsonify(ok=False, message="LLM 客户端不可用")
    try:
        config, err = jd_parser.create_job_from_jd(client, name, jd)
        if err:
            return jsonify(ok=False, message=err)
        return jsonify(ok=True, config=config)
    except Exception as e:
        logger.exception(e)
        return jsonify(ok=False, message=str(e))


# =============================================================================
# API：抓取 + 评估
# =============================================================================

@app.route("/api/scrape/open_browser", methods=["POST"])
def api_open_browser():
    if not _is_local_request():
        return jsonify(ok=False, local_only=True,
                       message="抓取功能仅限在运行本程序的电脑上操作。其他电脑可查看结果，但无法控制抓取浏览器。")
    data = request.get_json() or {}
    job_name = data.get("job_name")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name")
    state = pipeline.get_state(job_name)
    if state.get("running") and state.get("phase") == "scraping":
        return jsonify(ok=False,
                       message="抓取任务正在进行中，无法重新打开浏览器。请等待当前抓取完成或点击「中止抓取」。")
    return _open_browser_server(job_name)


@app.route("/api/scrape/start", methods=["POST"])
def api_scrape_start():
    if not _is_local_request():
        return jsonify(ok=False, local_only=True,
                       message="抓取功能仅限在运行本程序的电脑上操作。")
    data = request.get_json() or {}
    job_name = data.get("job_name")
    target_count = data.get("target_count")
    auto_evaluate = data.get("auto_evaluate", True)

    if not job_name:
        return jsonify(ok=False, message="缺少 job_name")
    driver = _get_driver()
    if not driver:
        return jsonify(ok=False, message="浏览器尚未打开。请先点击「打开浏览器」。")

    state = pipeline.get_state(job_name)
    if state.get("running"):
        return jsonify(ok=False, message="该岗位已有任务在运行")

    if target_count is None or str(target_count).strip() == "":
        job = database.get_job(job_name)
        cfg = json.loads(job["config_json"]) if job else {}
        target_count = int(cfg.get("target_count", 30))
    else:
        try:
            target_count = max(1, min(500, int(target_count)))
        except Exception:
            target_count = 30

    pipeline.reset_log(job_name)
    pipeline.reset_state(job_name)
    t = threading.Thread(
        target=_run_full_flow,
        args=(job_name, target_count, auto_evaluate),
        daemon=True
    )
    t.start()
    return jsonify(ok=True,
                   message=f"任务已启动：抓取目标 {target_count} 份"
                           + ("，抓取完后自动评估" if auto_evaluate else ""))


@app.route("/api/scrape/close_browser", methods=["POST"])
def api_close_browser():
    """主动关闭 Selenium 浏览器并清理状态"""
    if not _is_local_request():
        return jsonify(ok=False, local_only=True,
                       message="抓取功能仅限在运行本程序的电脑上操作。")
    _quit_driver()
    return jsonify(ok=True, message="浏览器已关闭")


@app.route("/api/scrape/stop", methods=["POST"])
def api_stop():
    if not _is_local_request():
        return jsonify(ok=False, local_only=True,
                       message="抓取功能仅限在运行本程序的电脑上操作。")
    data = request.get_json() or {}
    job_name = data.get("job_name")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name")
    pipeline.request_stop(job_name)
    pipeline._log(job_name, "⏹ 用户请求停止")
    return jsonify(ok=True, message="已发送停止信号，任务将在当前操作完成后退出")


@app.route("/api/status")
def api_status():
    job_name = request.args.get("job_name", "")
    state = pipeline.get_state(job_name)
    log_tail = pipeline.read_log_tail(job_name, max_lines=120)
    return jsonify(
        running=state.get("running", False),
        phase=state.get("phase", "idle"),
        current=state.get("current", 0),
        total=state.get("total", 0),
        done=state.get("done", False),
        last_msg=state.get("last_msg", ""),
        log=log_tail,
        error=state.get("error"),
        stop_requested=state.get("stop_requested", False),
    )


# =============================================================================
# API：Worker 守护进程控制
# =============================================================================

@app.route("/api/worker/status/<path:job_name>")
def api_worker_status(job_name):
    """返回从磁盘+DB派生的真实状态，重启网页后仍准确。"""
    return jsonify(worker.get_job_status(job_name))


@app.route("/api/worker/pause", methods=["POST"])
def api_worker_pause():
    job_name = (request.get_json() or {}).get("job_name")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name"), 400
    worker.pause_job(job_name)
    return jsonify(ok=True, message=f"{job_name} 评估已暂停")


@app.route("/api/worker/resume", methods=["POST"])
def api_worker_resume():
    job_name = (request.get_json() or {}).get("job_name")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name"), 400
    worker.resume_job(job_name)
    return jsonify(ok=True, message=f"{job_name} 评估已恢复")


@app.route("/api/evaluate/start", methods=["POST"])
def api_evaluate_start():
    """仅评估：基于已有 HTML 文件，跳过抓取阶段"""
    data = request.get_json() or {}
    job_name = data.get("job_name")
    force = bool(data.get("force_reevaluate", False))
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name")

    state = pipeline.get_state(job_name)
    if state.get("running"):
        return jsonify(ok=False, message="该岗位评估已在进行中")

    pipeline.reset_log(job_name)
    pipeline.reset_state(job_name)
    pipeline.run_evaluation_in_thread(job_name, force_reevaluate=force)
    return jsonify(ok=True, message="评估已开始")


@app.route("/api/reevaluate", methods=["POST"])
def api_reevaluate():
    data = request.get_json() or {}
    job_name = data.get("job_name")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name")

    state = pipeline.get_state(job_name)
    if state.get("running"):
        return jsonify(ok=False, message="该岗位有任务在运行，请先停止")

    pipeline.reset_log(job_name)
    pipeline.reset_state(job_name)
    pipeline.run_evaluation_in_thread(job_name, force_reevaluate=True)
    return jsonify(ok=True, message="重新评估已开始")


@app.route("/api/scrape/queue_submit", methods=["POST"])
def api_queue_submit():
    """任何局域网用户均可提交抓取任务；服务器串行执行"""
    data = request.get_json() or {}
    job_name = data.get("job_name")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name")
    if not database.get_job(job_name):
        return jsonify(ok=False, message=f"岗位「{job_name}」不存在")

    try:
        target_count = max(1, min(500, int(data.get("target_count", 30))))
    except Exception:
        target_count = 30
    auto_evaluate = bool(data.get("auto_evaluate", True))

    # 同岗位不重复入队
    with _queue_lock:
        if any(t["job_name"] == job_name for t in _scrape_queue):
            return jsonify(ok=False,
                           message=f"「{job_name}」已在队列中，无需重复提交")

    # 浏览器必须已由本机管理员打开
    if not _get_driver():
        return jsonify(ok=False,
                       message="浏览器尚未打开。请先由服务器管理员点击「打开浏览器」并登录 51job。")

    # 当前岗位有任务正在运行时，加入队列等待
    task = {
        "job_name":      job_name,
        "target_count":  target_count,
        "auto_evaluate": auto_evaluate,
        "submitted_by":  request.remote_addr,
        "submitted_at":  time.strftime("%H:%M:%S"),
    }
    with _queue_lock:
        _scrape_queue.append(task)
    pos = len(_get_queue_snapshot())
    return jsonify(ok=True,
                   message=f"任务已加入队列（当前第 {pos} 位），目标 {target_count} 份",
                   queue_position=pos)


@app.route("/api/scrape/queue_status")
def api_queue_status():
    """返回当前抓取队列（所有人可查看）"""
    snapshot = _get_queue_snapshot()
    # 同时返回当前正在运行的状态
    running_jobs = []
    for jn in {t["job_name"] for t in snapshot}:
        pass  # 队列里的都是等待中的
    return jsonify(ok=True, queue=snapshot, pending_count=len(snapshot))


@app.route("/api/scrape/queue_cancel", methods=["POST"])
def api_queue_cancel():
    """本机管理员可取消队列中的某个待执行任务"""
    if not _is_local_request():
        return jsonify(ok=False, message="仅服务器管理员可取消队列任务")
    job_name = (request.get_json() or {}).get("job_name")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name")
    with _queue_lock:
        before = len(_scrape_queue)
        new_q = collections.deque(
            t for t in _scrape_queue if t["job_name"] != job_name)
        _scrape_queue.clear()
        _scrape_queue.extend(new_q)
        removed = before - len(_scrape_queue)
    return jsonify(ok=True, removed=removed,
                   message=f"已从队列中移除 {removed} 个「{job_name}」任务")


# =============================================================================
# 人才池 API
# =============================================================================

@app.route("/talent_pool")
def talent_pool_page():
    jobs = database.list_jobs()
    return render_template("talent_pool.html", jobs=jobs)


@app.route("/api/talent_pool/add", methods=["POST"])
def api_pool_add():
    data = request.get_json() or {}
    eval_id = data.get("evaluation_id")
    recontact_date = data.get("recontact_date", "")
    reason_tag = data.get("reason_tag", "")
    note = data.get("note", "")
    if not eval_id or not recontact_date:
        return jsonify(ok=False, message="缺少 evaluation_id 或 recontact_date")
    pool_id = database.upsert_talent_pool(int(eval_id), recontact_date,
                                          reason_tag, note)
    return jsonify(ok=True, pool_id=pool_id)


@app.route("/api/talent_pool/remove", methods=["POST"])
def api_pool_remove():
    eval_id = (request.get_json() or {}).get("evaluation_id")
    if not eval_id:
        return jsonify(ok=False, message="缺少 evaluation_id")
    database.remove_from_talent_pool(int(eval_id))
    return jsonify(ok=True)


@app.route("/api/talent_pool/entry/<int:eval_id>")
def api_pool_entry(eval_id):
    entry = database.get_talent_pool_entry(eval_id)
    return jsonify(ok=True, entry=entry)


@app.route("/api/talent_pool/status", methods=["POST"])
def api_pool_status_update():
    data = request.get_json() or {}
    eval_id = data.get("evaluation_id")
    status = data.get("status")
    if not eval_id or not status:
        return jsonify(ok=False, message="缺少参数")
    database.update_talent_pool_status(int(eval_id), status)
    return jsonify(ok=True)


@app.route("/api/talent_pool/list")
def api_pool_list():
    job_name = request.args.get("job_name") or None
    status = request.args.get("status", "active")
    entries = database.list_talent_pool(job_name=job_name, status=status)
    return jsonify(ok=True, entries=entries, count=len(entries))


@app.route("/api/talent_pool/due")
def api_pool_due():
    days = int(request.args.get("days", 0))
    entries = database.list_due_talent_pool(days_ahead=days)
    return jsonify(ok=True, entries=entries, count=len(entries))


@app.route("/api/job/sessions_legacy")
def api_job_sessions():
    """（旧接口保留）返回该岗位所有评估批次及统计"""
    job_name = request.args.get("job_name", "")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name"), 400
    sessions = database.list_sessions_for_job(job_name)
    return jsonify(ok=True, sessions=sessions)


@app.route("/api/job/scrape_session")
def api_scrape_session():
    """返回当前岗位最近一次抓取批次信息；内存优先，重启后自动从 DB 恢复"""
    job_name = request.args.get("job_name", "")
    session = _scrape_sessions.get(job_name)
    if session:
        return jsonify(ok=True, has_session=True, count=len(session))

    # 内存为空（重启/首次）：从 evaluations 表恢复最近批次
    try:
        sessions = database.list_sessions_for_job(job_name)
        if sessions:
            latest_sid = sessions[0].get("scrape_session_id")
            if latest_sid:
                resume_ids = database.get_resume_ids_for_session(job_name, latest_sid)
                if resume_ids:
                    _scrape_sessions[job_name] = resume_ids   # 重新写回内存
                    return jsonify(ok=True, has_session=True, count=len(resume_ids))
    except Exception as e:
        logger.warning(f"恢复批次失败: {e}")

    return jsonify(ok=True, has_session=False, count=0)


@app.route("/api/job/scrape_session/clear", methods=["POST"])
def api_scrape_session_clear():
    job_name = (request.get_json() or {}).get("job_name", "")
    _scrape_sessions.pop(job_name, None)
    return jsonify(ok=True)


@app.route("/api/job/batch_history")
def api_batch_history():
    """返回某岗位最近 10 次批次历史（含本批次实时统计）。"""
    job_name = request.args.get("job_name", "")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name"), 400
    sessions = database.list_scrape_sessions(job_name, limit=10)

    # 最近一次批次的实时进度（覆盖 DB 数据，防止抓取中途数据陈旧）
    current_sid = None
    mem_session = _scrape_sessions.get(job_name)
    if mem_session:
        # 从 _session_meta 找到与该 job 匹配的最新 session_id
        with _agent_lock:
            for sid, meta in _session_meta.items():
                if meta.get("job_name") == job_name:
                    current_sid = sid
                    break

    result = []
    for s in sessions:
        sid = s["id"]
        if sid == current_sid:
            # 实时重算
            live = database.get_batch_stats(job_name, sid)
            s.update(live)
        result.append(s)

    return jsonify(ok=True, sessions=result, current_session_id=current_sid)


@app.route("/api/job/reeval_prefilter_batch", methods=["POST"])
def api_reeval_prefilter_batch():
    """
    清除指定批次的预筛拒绝记录并重新评估那批简历。
    输入: {job_name, session_id}
    输出: {ok, message}
    """
    data = request.get_json() or {}
    job_name = data.get("job_name", "")
    session_id = data.get("session_id", "")
    if not job_name or not session_id:
        return jsonify(ok=False, message="缺少 job_name 或 session_id"), 400
    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message="岗位不存在"), 404
    if pipeline.get_state(job_name).get("running"):
        return jsonify(ok=False, message="评估正在进行中，请等待完成"), 409
    resume_ids = database.get_prefilter_rejects_by_session(job["id"], session_id)
    if not resume_ids:
        return jsonify(ok=False, message="该批次无预筛拒绝记录（可能已清除）"), 404
    deleted = database.delete_prefilter_rejects_by_session(job["id"], session_id)
    pipeline.reset_log(job_name)
    pipeline.reset_state(job_name)
    pipeline.run_evaluation_for_ids_in_thread(job_name, resume_ids, force_reevaluate=False)
    return jsonify(ok=True,
                   message=f"已清除 {deleted} 条预筛记录，正在重新评估 {len(resume_ids)} 份简历")




@app.route("/api/job/errors")
def api_job_errors():
    """返回评估过程中的错误事件列表（前端轮询，since 参数支持增量获取）"""
    job_name = request.args.get("job_name", "")
    since = int(request.args.get("since", 0))
    events, total = pipeline.get_errors(job_name, since)
    return jsonify(ok=True, events=events, total=total)


@app.route("/api/data/clear_all", methods=["POST"])
def api_data_clear_all():
    """全量清除：候选人、评估、预过滤、行为记录、HTML 工作区、向量库（岗位配置保留）"""
    data = request.get_json() or {}
    confirm = data.get("confirm", "")
    if confirm != "CONFIRM_DELETE":
        return jsonify(ok=False, message="需要确认码 CONFIRM_DELETE")

    result = {}
    try:
        with database.get_conn() as conn:
            conn.execute("DELETE FROM talent_pool")
            conn.execute("DELETE FROM preference_signals")
            conn.execute("DELETE FROM preference_rules")
            conn.execute("DELETE FROM correction_signals")
            conn.execute("DELETE FROM job_criteria_notes")
            conn.execute("DELETE FROM outcomes")
            conn.execute("DELETE FROM rejection_tags")
            conn.execute("DELETE FROM prefilter_rejects")
            conn.execute("DELETE FROM evaluations")
            conn.execute("DELETE FROM candidates")
        result["db"] = "ok"
    except Exception as e:
        result["db"] = f"失败: {e}"

    # 删除所有工作区 HTML 文件
    ws_dirs = [
        d for d in os.listdir(ROOT)
        if d.startswith("工作区_") and os.path.isdir(os.path.join(ROOT, d))
    ]
    for ws in ws_dirs:
        try:
            shutil.rmtree(os.path.join(ROOT, ws))
            result[ws] = "deleted"
        except Exception as e:
            result[ws] = str(e)

    # 删除 Chroma 向量库目录（已弃用，保留清理逻辑）
    try:
        chroma_dir = os.path.join(ROOT, "data", "chroma")
        if os.path.exists(chroma_dir):
            shutil.rmtree(chroma_dir)
        result["vector"] = "ok"
    except Exception as e:
        result["vector"] = str(e)

    # 清理内存中的会话记录
    _scrape_sessions.clear()
    with _agent_lock:
        _session_meta.clear()

    # 清理错误缓冲
    for jn in list(pipeline._errors.keys()):
        pipeline.clear_errors(jn)

    logger.info(f"全量数据清除完成: {result}")
    return jsonify(ok=True, result=result)


# =============================================================================
# 分布式代理 API
# =============================================================================

@app.route("/api/agent/poll")
def api_agent_poll():
    """代理轮询：心跳 + 领取待执行命令。long_poll=1 时保持连接最多 25 秒等命令。"""
    did = request.args.get("device_id", "")
    if not did:
        return jsonify(ok=False)
    long_poll = request.args.get("long_poll", "0") == "1"

    def _update_and_drain():
        with _agent_lock:
            d = _devices.setdefault(did, {})
            d.update({
                "name": request.args.get("device_name", did),
                "last_seen": time.time(),
                "browser_open": request.args.get("browser_open", "0") == "1",
                "status": request.args.get("status", "idle"),
                "current_job": request.args.get("current_job", "") or None,
            })
            cmds = []
            if did in _device_commands:
                while _device_commands[did]:
                    cmds.append(_device_commands[did].popleft())
        return cmds

    cmds = _update_and_drain()
    if cmds or not long_poll:
        return jsonify(ok=True, commands=cmds)

    # 长轮询：挂起最多 25 秒，有命令时立即返回
    event = threading.Event()
    with _agent_lock:
        _device_events[did] = event
    try:
        event.wait(timeout=25)
    finally:
        with _agent_lock:
            _device_events.pop(did, None)

    # 更新心跳时间并取命令
    with _agent_lock:
        if did in _devices:
            _devices[did]["last_seen"] = time.time()
        cmds = []
        if did in _device_commands:
            while _device_commands[did]:
                cmds.append(_device_commands[did].popleft())
    return jsonify(ok=True, commands=cmds)


@app.route("/api/agent/progress", methods=["POST"])
def api_agent_progress():
    """代理上报进度，同步到 pipeline 状态供驾驶舱轮询"""
    data = request.get_json() or {}
    did = data.get("device_id", "")
    job_name = data.get("job_name", "")
    phase = data.get("phase", "idle")
    current = int(data.get("current", 0))
    total = int(data.get("total", 0))
    msg = data.get("message", "")

    # 下载阶段在 _devices 中显示 "downloading"，完成后回到 "idle"
    _DONE_PHASES = {"done", "stopped", "error", "dl_done", "dl_error"}
    if phase == "downloading":
        device_status = "downloading"
    elif phase in _DONE_PHASES:
        device_status = "idle"
    else:
        device_status = phase

    with _agent_lock:
        d = _devices.setdefault(did, {})
        d.update({
            "name": data.get("device_name", did),
            "last_seen": time.time(),
            "browser_open": bool(data.get("browser_open", False)),
            "status": device_status,
            "current_job": job_name if phase == "scraping" else None,
        })

    # 扩展本机下载进度更新（漏洞3修复：服务器追踪任务状态供前端轮询）
    if phase in ("downloading", "dl_done", "dl_error"):
        batch_id = data.get("batch_id", "")
        failed   = int(data.get("failed", 0))
        with _ext_dl_lock:
            task = _ext_dl_tasks.get(did)
            if task and task.get("batch_id") == batch_id:
                task["done"]         = current
                task["failed"]       = failed
                task["current_name"] = msg
                if phase == "downloading" and msg.startswith("⚠"):
                    # 保存最新的单项失败原因，供最终错误展示用
                    task["last_item_error"] = msg
                elif phase == "dl_done":
                    task["state"]    = "done"
                elif phase == "dl_error":
                    task["state"]    = "error"
                    task["error_msg"] = msg

    if job_name and phase not in ("downloading", "dl_done", "dl_error"):
        running = phase == "scraping"
        pipeline._set_state(job_name, running=running,
                            phase=phase if running else "idle",
                            current=current, total=total,
                            last_msg=msg, done=(phase == "done"))
        if msg:
            pipeline._log(job_name, f"[{data.get('device_name', did)}] {msg}")
        # 抓取完成时结算批次统计
        if phase == "done":
            sid = data.get("session_id", "")
            if sid:
                try:
                    database.finish_scrape_session(sid)
                except Exception:
                    pass
    return jsonify(ok=True)


@app.route("/api/agent/list")
def api_agent_list():
    """列出所有在线/离线设备（含服务器本机作为内置设备）"""
    now = time.time()
    result = []
    # 服务器本机：始终在列表中（browser_open = WebDriver 已连接 或 进程正在运行）
    result.append({
        "device_id": SERVER_DEVICE_ID,
        "name": SERVER_DEVICE_NAME,
        "online": True,
        "browser_open": _browser_is_open(),
        "status": "scraping" if pipeline.get_state("").get("running") else "idle",
        "current_job": _driver_state.get("job_name"),
        "is_server": True,
    })
    with _agent_lock:
        for did, d in _devices.items():
            online = (now - d.get("last_seen", 0)) < DEVICE_TIMEOUT
            result.append({
                "device_id": did,
                "name": d.get("name", did),
                "online": online,
                "browser_open": d.get("browser_open", False),
                "status": d.get("status", "idle") if online else "offline",
                "current_job": d.get("current_job"),
                "is_server": False,
            })
    return jsonify(ok=True, devices=result)


@app.route("/api/agent/command", methods=["POST"])
def api_agent_command():
    """驾驶舱向指定设备发送命令（open_browser / scrape / stop / close_browser）"""
    data = request.get_json() or {}
    did = data.get("device_id", "")
    cmd = data.get("command", "")
    if not did or not cmd:
        return jsonify(ok=False, message="缺少 device_id 或 command")

    # 如果目标是服务器本机，直接走现有代码路径
    if did == SERVER_DEVICE_ID:
        job_name = data.get("job_name", "")
        if cmd == "open_browser":
            if not _is_local_request():
                return jsonify(ok=False, message="服务器本机浏览器只能从服务器本机触发")
            return api_scrape_open_browser_impl(job_name)
        elif cmd == "scrape":
            if not _is_local_request():
                return jsonify(ok=False, message="服务器本机浏览器只能从服务器本机触发")
            tc = int(data.get("target_count", 30))
            return api_scrape_start_impl(job_name, tc)
        elif cmd == "close_browser":
            _quit_driver()
            return jsonify(ok=True)
        elif cmd == "stop":
            pipeline.request_stop(job_name)
            return jsonify(ok=True)
        return jsonify(ok=False, message=f"未知命令: {cmd}")

    # 远程代理设备：加入命令队列并唤醒长轮询等待
    with _agent_lock:
        if did not in _device_commands:
            _device_commands[did] = collections.deque()
        _device_commands[did].append(data)
        ev = _device_events.get(did)
    if ev:
        ev.set()
    return jsonify(ok=True)


@app.route("/api/scrape/upload_html", methods=["POST"])
def api_upload_html():
    """代理上传简历 HTML；服务器保存文件，worker 自动触发评估"""
    data = request.get_json() or {}
    job_name  = data.get("job_name", "")
    resume_id = data.get("resume_id", "")
    html      = data.get("html", "")
    session_id = data.get("session_id", "")
    device_id  = data.get("device_id", "")
    device_name = data.get("device_name", device_id)
    name_hint  = (data.get("name_hint") or "未知").replace("/", "").replace("\\", "")[:30]
    index      = int(data.get("index", 1))

    if not (job_name and resume_id and html):
        return jsonify(ok=False, message="缺少必要字段")

    html_dir = pipeline._resolve_workspace_dir(job_name)
    os.makedirs(html_dir, exist_ok=True)
    fname = f"{index:03d}_{name_hint}_{resume_id}.html"
    fpath = os.path.join(html_dir, fname)
    gz_path = fpath + ".gz"

    if not os.path.exists(fpath) and not os.path.exists(gz_path):
        try:
            with open(fpath, "w", encoding="utf-8") as f:
                f.write(html)
        except Exception as e:
            return jsonify(ok=False, message=str(e))

    # 记录批次 session
    if session_id:
        _scrape_sessions.setdefault(job_name, set()).add(resume_id)
        with _agent_lock:
            if session_id not in _session_meta:
                _session_meta[session_id] = {"device_name": device_name, "job_name": job_name}
                # 首次见到该 session_id，写入 scrape_sessions 表
                try:
                    database.upsert_scrape_session(session_id, job_name, device_name)
                except Exception:
                    pass

    pipeline._log(job_name,
                  f"  ↑ [{device_name}] ({data.get('index')}/{data.get('total')}) {name_hint}")
    return jsonify(ok=True)


@app.route("/api/scrape/upload_file", methods=["POST"])
def api_upload_resume_file():
    """接收远程代理上传的简历 PDF/Word 文件（base64 编码）"""
    data      = request.get_json() or {}
    resume_id = data.get("resume_id", "").strip()
    filetype  = data.get("filetype", "").strip()   # "pdf" or "word"
    data_b64  = data.get("data_b64", "")
    if not (resume_id and filetype in ("pdf", "word") and data_b64):
        return jsonify(ok=False, message="缺少必要字段")
    try:
        file_bytes = base64.b64decode(data_b64)
        if len(file_bytes) < 200:
            return jsonify(ok=False, message="文件内容无效")
        dest = _resume_file_path(resume_id, filetype)
        with open(dest, "wb") as f:
            f.write(file_bytes)
        return jsonify(ok=True)
    except Exception as e:
        return jsonify(ok=False, message=str(e))


@app.route("/api/resume/download/<resume_id>/<filetype>")
def api_resume_download(resume_id, filetype):
    """向前端提供简历 PDF 或 Word 文件下载"""
    if not _RESUME_ID_RE.match(resume_id):
        return "无效ID", 400
    if filetype not in ("pdf", "word"):
        return "无效文件类型", 400
    path = _resume_file_path(resume_id, filetype)
    if not os.path.exists(path):
        return "文件不存在，请先抓取时下载或在驾驶舱重新触发下载", 404
    ext  = ".pdf" if filetype == "pdf" else ".docx"
    mime = ("application/pdf" if filetype == "pdf"
            else "application/vnd.openxmlformats-officedocument.wordprocessingml.document")
    return send_file(path, as_attachment=True,
                     download_name=f"{resume_id}{ext}", mimetype=mime)


@app.route("/api/resume/view/<resume_id>/pdf")
def api_resume_view_pdf(resume_id):
    """向处理台 iframe 提供 PDF 内嵌预览（非附件下载）"""
    if not _RESUME_ID_RE.match(resume_id):
        return "无效ID", 400
    path = _resume_file_path(resume_id, "pdf")
    if not os.path.exists(path):
        return "PDF 文件不存在", 404
    return send_file(path, mimetype="application/pdf")


@app.route("/api/batch_download", methods=["POST"])
def api_batch_download():
    """批量下载简历 PDF，打包成 ZIP 返回。
    Body: {items: [{resume_id, name}], job_name: str}
    仅下载本地已存在的文件，跳过未下载的。单文件时直接返回原文件不打包。
    """
    data = request.get_json() or {}
    items    = data.get("items", [])
    job_name = (data.get("job_name") or "岗位")[:50]

    if not items:
        return jsonify(ok=False, message="未选择任何候选人"), 400

    ext         = ".pdf"
    mime_single = "application/pdf"

    found = []
    for item in items[:200]:
        rid = str(item.get("resume_id", "")).strip()
        # 严格校验 resume_id，防止路径穿越（如 "../../secret"）
        if not _RESUME_ID_RE.match(rid):
            continue
        raw_name = (item.get("name") or rid).strip()
        # 清洗姓名：防止路径分隔符进入 ZIP entry（ZipSlip 防御）
        name = re.sub(r'[/\\:*?"<>|]', '_', raw_name)[:50] or rid[:50]
        date = re.sub(r'[^0-9\-]', '', str(item.get("date", ""))).strip()[:10]
        path = os.path.join(FILES_DIR, rid + ext)
        if os.path.exists(path):
            found.append((path, name, rid, date))

    if not found:
        return jsonify(ok=False, message="所选候选人均无本地文件，请先在驾驶舱触发下载"), 400

    def _entry_name(name, rid, date):
        named = _build_named_filename(date, job_name, name)
        return named if named else f"{name}_{rid[:8]}{ext}"

    if len(found) == 1:
        path, name, rid, date = found[0]
        filename = _entry_name(name, rid, date)
        return send_file(path, as_attachment=True, download_name=filename, mimetype=mime_single)

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        seen = {}
        for path, name, rid, date in found:
            arcname = _entry_name(name, rid, date)
            if arcname in seen:
                seen[arcname] += 1
                stem = arcname[:-4]
                arcname = f"{stem}_{seen[arcname]}{ext}"
            else:
                seen[arcname] = 1
            zf.write(path, arcname)
    buf.seek(0)

    safe_job = re.sub(r'[/\\:*?"<>|]', '_', str(job_name))[:30]
    zip_name = f"{safe_job}_简历_PDF.zip"
    return send_file(buf, mimetype="application/zip", as_attachment=True, download_name=zip_name)


# =============================================================================
# 后台批量下载 API（通过 Selenium 浏览器从51job下载到本地）
# =============================================================================

@app.route("/api/resume/bg_download", methods=["POST"])
def api_bg_download():
    """
    启动后台批量下载任务（仅 PDF）。
    Body: {items:[{resume_id,name}], job_name:str}
    最多同时运行一个下载任务；浏览器未打开或正在抓取时拒绝。
    """
    data     = request.get_json() or {}
    items    = data.get("items", [])
    job_name = (data.get("job_name") or "岗位")[:50]
    filetype = "pdf"

    # 校验并清洗 resume_id（防路径穿越 / URL 注入）；清洗 name（防 ZipSlip）
    validated = []
    for item in items[:200]:
        rid  = str(item.get("resume_id", "")).strip()
        if not _RESUME_ID_RE.match(rid):
            continue
        raw_name = str(item.get("name", rid))
        # 移除路径分隔符和 ZIP 危险字符，防止 ZipSlip 攻击
        name = re.sub(r'[/\\:*?"<>|]', '_', raw_name).strip()[:50] or rid[:50]
        # 只保留数字和连字符，防止日期字段注入路径穿越字符
        date = re.sub(r'[^0-9\-]', '', str(item.get("date", ""))).strip()[:10]
        validated.append({"resume_id": rid, "name": name, "date": date})

    if not validated:
        return jsonify(ok=False, message="无合法的候选人（resume_id 格式错误）"), 400

    # 浏览器必须已打开
    if not _browser_is_open():
        return jsonify(
            ok=False,
            message="浏览器未打开。请先在驾驶舱点击「打开浏览器」并登录51job，再使用后台下载。"
        ), 400

    # 抓取进行中时拒绝（避免两个任务争用浏览器）
    with _driver_lock:
        active_job = _driver_state.get("job_name")
    if active_job and pipeline.get_state(active_job).get("running"):
        return jsonify(
            ok=False,
            message=f"正在抓取「{active_job}」，请等待抓取完成后再使用后台下载。"
        ), 409

    # 原子性占用槽位（锁内检查+设置，防并发竞争）
    with _dl_lock:
        if _dl_status["state"] == "running":
            return jsonify(
                ok=False,
                message="已有下载任务正在运行，请等待完成或点击取消后重试。"
            ), 409
        _dl_status["state"] = "running"               # 抢占槽位

    _dl_cancel_flag.clear()
    t = threading.Thread(
        target=_bg_download_worker,
        args=(validated, filetype, job_name),
        daemon=True,
        name="bg-dl-worker",
    )
    t.start()

    return jsonify(ok=True, count=len(validated),
                   message=f"已开始后台下载 {len(validated)} 份简历")


@app.route("/api/resume/bg_status")
def api_bg_status():
    """返回后台下载进度（前端每 2 秒轮询一次）。
    运行中不返回 results 列表（避免大量轮询浪费带宽），完成后才附带。
    """
    with _dl_lock:
        snap = {k: v for k, v in _dl_status.items() if k != "results"}
        # 完成/取消/错误时才附带结果列表（最多200条）
        if _dl_status["state"] in ("done", "cancelled", "error"):
            snap["results"] = _dl_status["results"][:200]
    return jsonify(ok=True, **snap)


@app.route("/api/resume/bg_cancel", methods=["POST"])
def api_bg_cancel():
    """取消正在进行的后台下载（信号式，worker 在下一个 resume 检查时退出）"""
    _dl_cancel_flag.set()
    with _dl_lock:
        if _dl_status["state"] == "running":
            _dl_status["state"]        = "cancelled"
            _dl_status["completed_at"] = time.time()
    return jsonify(ok=True, message="已发送取消信号")


@app.route("/api/resume/ext_download", methods=["POST"])
def api_ext_download():
    """
    向指定扩展设备发送批量本机下载命令。
    每台设备独立执行，互不冲突；文件保存到该设备的本机下载目录。
    Body: {items:[{resume_id,name}], job_name:str, device_id:str}
    """
    data      = request.get_json() or {}
    device_id = str(data.get("device_id", "")).strip()
    job_name  = (data.get("job_name") or "岗位")[:50]
    items     = data.get("items", [])

    if not device_id:
        return jsonify(ok=False, message="缺少 device_id"), 400

    # 漏洞5修复：拒绝服务器本机（服务器本机用 bg_download）
    if device_id == SERVER_DEVICE_ID:
        return jsonify(ok=False, message="服务器本机请使用「下载到服务器」"), 400

    # 检查设备在线（漏洞3：离线设备发命令无意义）
    now = time.time()
    with _agent_lock:
        d = _devices.get(device_id)
        if not d or (now - d.get("last_seen", 0)) > DEVICE_TIMEOUT:
            return jsonify(ok=False, message="该设备当前不在线，请确认扩展已打开并连接服务器"), 400

    # 验证 items（与 bg_download 相同逻辑，防路径穿越）
    validated = []
    for item in items[:200]:
        rid      = str(item.get("resume_id", "")).strip()
        if not _RESUME_ID_RE.match(rid):
            continue
        raw_name = str(item.get("name", rid))
        name     = re.sub(r'[/\\:*?"<>|]', '_', raw_name).strip()[:50] or rid[:50]
        date     = re.sub(r'[^0-9\-]', '', str(item.get("date", ""))).strip()[:10]
        validated.append({"resume_id": rid, "name": name, "date": date})

    if not validated:
        return jsonify(ok=False, message="无合法的候选人（resume_id 格式错误）"), 400

    # 漏洞4修复：若该设备已有运行中任务则拒绝（409）
    batch_id = f"{int(time.time()*1000):x}{random.randint(0, 0xffffff):06x}"[:16]
    with _ext_dl_lock:
        task = _ext_dl_tasks.get(device_id, {})
        if task.get("state") == "running":
            return jsonify(ok=False,
                           message="该设备已有下载任务正在运行，请等待完成后重试"), 409
        _ext_dl_tasks[device_id] = {
            "state": "running", "done": 0, "failed": 0,
            "total": len(validated), "current_name": "",
            "job_name": job_name, "batch_id": batch_id,
            "error_msg": "", "last_item_error": "", "started_at": time.time(),
        }

    # 推送 batch_download 命令到设备的长轮询队列
    cmd = {
        "command":  "batch_download",
        "items":    validated,
        "job_name": job_name,
        "batch_id": batch_id,
    }
    with _agent_lock:
        if device_id not in _device_commands:
            _device_commands[device_id] = collections.deque()
        _device_commands[device_id].append(cmd)
        ev = _device_events.get(device_id)
        if ev:
            ev.set()

    return jsonify(ok=True, count=len(validated), batch_id=batch_id,
                   message=f"已向设备发送下载命令，共 {len(validated)} 份")


@app.route("/api/resume/ext_dl_status")
def api_ext_dl_status():
    """
    查询指定设备的本机下载任务进度（前端轮询）。
    漏洞3修复：若设备离线且任务仍在 running，自动标记为 disconnected。
    """
    device_id = request.args.get("device_id", "").strip()
    if not device_id:
        return jsonify(ok=False, message="缺少 device_id"), 400

    with _ext_dl_lock:
        task = dict(_ext_dl_tasks.get(device_id, {}))

    if not task:
        return jsonify(ok=True, state="idle")

    # 漏洞3修复：任务 running 但设备已离线 → 标记 disconnected
    if task.get("state") == "running":
        now = time.time()
        with _agent_lock:
            d = _devices.get(device_id, {})
            online = (now - d.get("last_seen", 0)) < DEVICE_TIMEOUT
        if not online:
            with _ext_dl_lock:
                if _ext_dl_tasks.get(device_id, {}).get("state") == "running":
                    _ext_dl_tasks[device_id]["state"] = "disconnected"
            task["state"] = "disconnected"

    return jsonify(ok=True, **task)


@app.route("/api/resume/file_info", methods=["POST"])
def api_resume_file_info():
    """批量查询指定 resume_id 列表的文件存在情况"""
    raw = (request.get_json() or {}).get("resume_ids", [])
    if not raw:
        return jsonify(ok=True, pdf=[], word=[])
    # 过滤非法 ID，防止路径探测
    resume_ids = [str(r) for r in raw[:500] if _RESUME_ID_RE.match(str(r))]
    if not resume_ids:
        return jsonify(ok=True, pdf=[], word=[])
    has_pdf, has_word = _check_resume_files(resume_ids)
    return jsonify(ok=True, pdf=list(has_pdf), word=list(has_word))


@app.route("/api/resume/download_online", methods=["POST"])
def api_resume_download_online():
    """按需下载：让有浏览器的设备访问51job页面实时下载文件，再传回服务器供前端取走。
    立即返回 {ok, queued/ready}，前端通过轮询 /api/resume/file_info 等待完成。
    """
    data = request.get_json() or {}
    resume_id = data.get("resume_id", "").strip()
    filetype  = data.get("filetype", "pdf")
    if not resume_id or filetype not in ("pdf", "word"):
        return jsonify(ok=False, message="参数错误"), 400

    # 文件已存在，直接告知就绪
    fpath = _resume_file_path(resume_id, filetype)
    if os.path.exists(fpath):
        return jsonify(ok=True, ready=True)

    # 优先服务器本机浏览器，其次找在线远程设备
    server_driver = _get_driver()
    remote_did = None
    if not server_driver:
        now = time.time()
        with _agent_lock:
            for did, d in _devices.items():
                if (now - d.get("last_seen", 0)) < DEVICE_TIMEOUT and d.get("browser_open"):
                    remote_did = did
                    break

    if not server_driver and not remote_did:
        return jsonify(ok=False, message="没有设备打开了浏览器，请先点击「打开浏览器」并登录51job"), 400

    if server_driver:
        # 服务器本机：后台线程执行下载
        if pipeline.get_state("").get("running"):
            return jsonify(ok=False, message="服务器正在抓取中，请稍后再试"), 409

        def _server_dl():
            hdrv = None
            try:
                # localStorage 是域名隔离的，必须先导航到 ehire.51job.com 才能读取
                if "ehire.51job.com" not in server_driver.current_url:
                    server_driver.get("https://ehire.51job.com")
                    time.sleep(2.5)
                cookies, local_storage, session_storage = _extract_browser_auth(server_driver)
                hdrv = _create_headless_driver(cookies, local_storage, session_storage)
                url = (f"https://ehire.51job.com/Revision/talent/resume/detail"
                       f"?resumeId={resume_id}")
                hdrv.get(url)
                time.sleep(3)
                final_url = hdrv.current_url
                if "resumeId" not in final_url or "detail" not in final_url:
                    raise RuntimeError(f"被重定向（{final_url}），请重新登录 51job")
                cdp_result = hdrv.execute_cdp_cmd("Page.printToPDF", {
                    "printBackground": True, "format": "A4",
                    "marginTop": 0.5, "marginBottom": 0.5,
                    "marginLeft": 0.4, "marginRight": 0.4, "scale": 1.0,
                })
                pdf_bytes = base64.b64decode(cdp_result.get("data", ""))
                if pdf_bytes and pdf_bytes[:4] == b"%PDF":
                    with open(os.path.join(FILES_DIR, resume_id + ".pdf"), "wb") as f:
                        f.write(pdf_bytes)
                    logger.info(f"[按需下载] PDF 已保存: {resume_id}")
                else:
                    raise RuntimeError("Page.printToPDF 返回数据为空或不是有效PDF")
            except Exception as e:
                logger.warning(f"按需下载失败 ({resume_id}): {e}")
            finally:
                if hdrv:
                    try: hdrv.quit()
                    except Exception: pass

        threading.Thread(target=_server_dl, daemon=True).start()
        return jsonify(ok=True, queued=True)

    else:
        # 远程设备：下发 download_file 命令
        cmd = {"command": "download_file", "resume_id": resume_id, "filetype": filetype}
        with _agent_lock:
            if remote_did not in _device_commands:
                _device_commands[remote_did] = collections.deque()
            _device_commands[remote_did].append(cmd)
            ev = _device_events.get(remote_did)
        if ev:
            ev.set()
        return jsonify(ok=True, queued=True)


@app.route("/api/job/sessions")
def api_job_sessions_meta():
    """返回岗位的所有评估批次，附带抓取设备名（供驾驶舱批次视图使用）"""
    job_name = request.args.get("job_name", "")
    sessions = database.list_sessions_for_job(job_name)
    with _agent_lock:
        for s in sessions:
            sid = s.get("scrape_session_id") or ""
            meta = _session_meta.get(sid, {})
            s["device_name"] = meta.get("device_name", "")
    return jsonify(ok=True, sessions=sessions)


def api_scrape_open_browser_impl(job_name):
    """服务器本机打开浏览器（供 agent/command 路由）"""
    return _open_browser_server(job_name)


def api_scrape_start_impl(job_name, target_count):
    """服务器本机启动抓取（供 agent/command 路由）"""
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name")
    if not _get_driver():
        return jsonify(ok=False, message="浏览器尚未打开，请先打开浏览器")
    if pipeline.get_state(job_name).get("running"):
        return jsonify(ok=False, message="已有任务在运行")
    t = threading.Thread(
        target=_run_full_flow,
        args=(job_name, target_count, True, True),
        daemon=True,
    )
    t.start()
    return jsonify(ok=True, message=f"抓取已启动，目标 {target_count} 份")


@app.route("/api/evaluate/batch", methods=["POST"])
def api_evaluate_batch():
    """仅评估/重评本次抓取批次中的简历"""
    data = request.get_json() or {}
    job_name = data.get("job_name", "")
    force = bool(data.get("force_reevaluate", False))
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name")
    session = _scrape_sessions.get(job_name)
    if not session:
        return jsonify(ok=False, message="当前没有活跃的抓取批次，请先抓取简历")
    if pipeline.get_state(job_name).get("running"):
        return jsonify(ok=False, message="评估正在进行中，请等待完成")
    pipeline.reset_log(job_name)
    pipeline.reset_state(job_name)
    pipeline.run_evaluation_for_ids_in_thread(job_name, session, force_reevaluate=force)
    return jsonify(ok=True,
                   message=f"批次{'重评' if force else '评估'}已启动（共 {len(session)} 份）")


@app.route("/api/job/cleanup/candidates")
def api_cleanup_candidates_list():
    """清理台：获取该岗位所有 LLM 评估候选人（含 HR 操作状态）"""
    job_name = request.args.get("job_name", "")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name"), 400
    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message="岗位不存在"), 404
    with database.get_conn() as conn:
        rows = conn.execute("""
            SELECT e.id AS evaluation_id, e.candidate_id, e.verdict,
                   c.resume_id, c.name, c.age, c.first_degree,
                   (SELECT action FROM outcomes o
                    WHERE o.evaluation_id = e.id
                    ORDER BY o.id DESC LIMIT 1) AS latest_action
            FROM evaluations e
            JOIN candidates c ON e.candidate_id = c.id
            WHERE e.job_id = ?
            ORDER BY
              CASE e.verdict WHEN '排除' THEN 0 WHEN '黄色' THEN 1
                             WHEN '蓝色' THEN 2 WHEN '深绿' THEN 3 ELSE 4 END,
              e.id DESC
        """, (job["id"],)).fetchall()
    return jsonify(ok=True, candidates=[dict(r) for r in rows])


@app.route("/api/job/cleanup/delete", methods=["POST"])
def api_cleanup_delete():
    """清理台：删除当前岗位的评估记录；仅当候选人在所有岗位均无评估时才删除档案和文件"""
    data = request.get_json() or {}
    job_name = data.get("job_name", "")
    items = data.get("items", [])
    if not job_name or not items:
        return jsonify(ok=False, message="参数缺失"), 400

    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message="岗位不存在"), 404
    job_id = job["id"]

    html_dir = pipeline._resolve_workspace_dir(job_name)
    deleted_evals = 0
    deleted_candidates = 0
    deleted_files = 0

    for item in items:
        resume_id   = item.get("resume_id")
        eval_id_raw = item.get("evaluation_id")

        try:
            eval_id = int(eval_id_raw)
        except (TypeError, ValueError):
            logger.warning(f"无效的 evaluation_id: {eval_id_raw!r}，跳过")
            continue

        # 1. 删除本岗位的评估记录及所有 FK 子记录
        database.delete_evaluation(eval_id)
        deleted_evals += 1

        # 2. 删除本岗位的预筛拒绝记录（prefilter_rejects 没有 evaluation_id，需单独清理）
        if resume_id:
            with database.get_conn() as conn:
                conn.execute(
                    "DELETE FROM prefilter_rejects WHERE resume_id=? AND job_id=?",
                    (resume_id, job_id)
                )

        # 3. 删除本岗位工作区 HTML（精确匹配 resume_id，避免子串误删）
        if resume_id and html_dir and os.path.exists(html_dir):
            for fname in os.listdir(html_dir):
                frid, _ = pipeline._parse_html_filename(fname)
                if frid == resume_id:
                    try:
                        os.remove(os.path.join(html_dir, fname))
                        deleted_files += 1
                    except Exception as e:
                        logger.warning(f"删除工作区文件失败 {fname}: {e}")

        # 4. 检查候选人是否还有其他岗位的评估；有则保留档案和文件
        if not resume_id:
            continue
        candidate = database.get_candidate_by_resume_id(resume_id)
        if not candidate:
            continue
        if database.count_evaluations_for_candidate(candidate["id"]) > 0:
            continue

        # 5. 无任何剩余评估，彻底删除候选人档案和所有岗位的预筛记录
        with database.get_conn() as conn:
            conn.execute("DELETE FROM prefilter_rejects WHERE resume_id=?", (resume_id,))
            conn.execute("DELETE FROM candidates WHERE id=?", (candidate["id"],))
        deleted_candidates += 1

        # 6. 删除 FILES_DIR 中的原始简历和文字 HTML（精确路径，不扫全目录）
        for ext in (".pdf", ".docx", ".doc"):
            p = os.path.join(FILES_DIR, resume_id + ext)
            if os.path.exists(p):
                try:
                    os.remove(p)
                    deleted_files += 1
                except Exception as e:
                    logger.warning(f"删除简历文件失败 {p}: {e}")
        text_html = os.path.join(FILES_DIR, resume_id + "_text.html")
        if os.path.exists(text_html):
            try:
                os.remove(text_html)
                deleted_files += 1
            except Exception as e:
                logger.warning(f"删除文字 HTML 失败 {text_html}: {e}")

    msg = f"已删除 {deleted_evals} 条评估记录"
    if deleted_candidates:
        msg += f"，彻底移除 {deleted_candidates} 位候选人"
    if deleted_files:
        msg += f"，{deleted_files} 份文件"
    return jsonify(ok=True, deleted_evals=deleted_evals,
                   deleted_candidates=deleted_candidates,
                   deleted_files=deleted_files, message=msg)


@app.route("/api/job/reset", methods=["POST"])
def api_job_reset():
    """完全重置：清空该岗位所有评估记录 + 删除工作区所有简历文件"""
    data = request.get_json() or {}
    job_name = data.get("job_name")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name"), 400

    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message="岗位不存在"), 404

    state = pipeline.get_state(job_name)
    if state.get("running"):
        return jsonify(ok=False, message="有任务正在运行，请先停止"), 400

    job_id = job["id"]

    # 1. 删除评估记录（含 prefilter_rejects / outcomes / rejection_tags）
    database.delete_evaluations_for_job(job_id)

    # 2. 删除工作区 HTML 文件
    html_dir = pipeline._resolve_workspace_dir(job_name)
    deleted_files = 0
    if html_dir and os.path.exists(html_dir):
        for f in os.listdir(html_dir):
            if f.endswith(".html") or f.endswith(".html.gz"):
                try:
                    os.remove(os.path.join(html_dir, f))
                    deleted_files += 1
                except Exception as e:
                    logger.warning(f"删除文件失败 {f}: {e}")

    pipeline.reset_log(job_name)
    pipeline.reset_state(job_name)
    return jsonify(ok=True,
                   message=f"已清空所有评估记录，删除 {deleted_files} 份简历文件")


# =============================================================================
# 抓取-评估完整流（后台线程）
# =============================================================================

def _run_full_flow(job_name, target_count, auto_evaluate,
                   close_browser_after=True):
    """执行完整抓取流程。
    close_browser_after=True  → 流程结束后关闭浏览器（本机直接启动时的默认行为）
    close_browser_after=False → 保留浏览器供下一个队列任务复用（队列模式）
    """
    try:
        success = _do_scrape(job_name, target_count)

        if pipeline.is_stop_requested(job_name):
            pipeline._set_state(job_name, running=False, done=True,
                                phase="done", last_msg="已停止")
            return

        if not success:
            pipeline._set_state(job_name, running=False, done=True,
                                phase="done")
            return

        if auto_evaluate:
            pipeline._log(job_name, "▶ 抓取完成，开始 AI 评估")
            pipeline.run_evaluation(job_name)
        else:
            pipeline._set_state(job_name, running=False, done=True,
                                phase="done", last_msg="抓取完成")
    finally:
        if close_browser_after:
            _quit_driver()


def _do_scrape(job_name, target_count):
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    import driver_utils

    driver = _get_driver()
    if not driver:
        pipeline._set_state(job_name, running=False, done=True,
                            error="浏览器未打开")
        return False

    pipeline._set_state(job_name, running=True, phase="scraping",
                        current=0, total=target_count,
                        done=False, error=None, stop_requested=False)
    pipeline._log(job_name, f"开始抓取，目标 {target_count} 份")

    workspace = os.path.join(ROOT, f"工作区_{pipeline.safe_filename(job_name)}")
    html_dir = os.path.join(workspace, "简历原始文件")
    os.makedirs(html_dir, exist_ok=True)

    try:
        task_pool = []
        seen_ids = set()
        no_progress = 0

        while len(task_pool) < target_count and no_progress < 10:
            if pipeline.is_stop_requested(job_name):
                pipeline._log(job_name, "⏹ 抓取阶段被用户停止")
                return False

            cards = driver.find_elements(
                By.XPATH, "//div[contains(@class,'resume-card')]")
            found_new = False
            for card in cards:
                rid = driver_utils.extract_resume_id(card)
                if rid and rid not in seen_ids:
                    name = driver_utils.get_clean_name(card)
                    task_pool.append({"id": rid, "name": name})
                    seen_ids.add(rid)
                    found_new = True
                    if len(task_pool) >= target_count:
                        break

            pipeline._log(job_name,
                          f"  扫描到候选人 {len(task_pool)}/{target_count}")
            pipeline._set_state(job_name, current=len(task_pool))

            if len(task_pool) >= target_count:
                break
            prev_card_count = len(cards)
            try:
                if cards:
                    driver.execute_script("""
                        var last = arguments[0];
                        last.scrollIntoView({block: 'end', behavior: 'instant'});
                        var el = last.parentElement;
                        while (el && el !== document.body) {
                            var style = window.getComputedStyle(el);
                            if (style.overflowY === 'auto' || style.overflowY === 'scroll') {
                                el.scrollTop = el.scrollHeight;
                                return;
                            }
                            el = el.parentElement;
                        }
                        window.scrollTo(0, document.body.scrollHeight);
                    """, cards[-1])
            except Exception:
                pass
            try:
                WebDriverWait(driver, 5).until(
                    lambda d: len(d.find_elements(
                        By.XPATH, "//div[contains(@class,'resume-card')]")) > prev_card_count
                )
            except Exception:
                pass
            no_progress = no_progress + 1 if not found_new else 0

        pipeline._log(job_name, f"开始保存 {len(task_pool)} 份简历详情页")
        pipeline._set_state(job_name, total=len(task_pool), current=0)

        for i, task in enumerate(task_pool):
            if pipeline.is_stop_requested(job_name):
                pipeline._log(job_name, "⏹ 详情页保存阶段被用户停止")
                return False

            file_name = f"{i+1:03d}_{task['name']}_{task['id']}.html"
            file_path = os.path.join(html_dir, file_name)
            gz_path = file_path + ".gz"
            pipeline._set_state(job_name, current=i+1)
            if os.path.exists(file_path) or os.path.exists(gz_path):
                continue
            pipeline._log(job_name,
                          f"  保存 ({i+1}/{len(task_pool)}): {task['name']}")
            try:
                driver.get(f"https://ehire.51job.com/Revision/talent/resume/detail?resumeId={task['id']}")
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.ID, "work")))
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(driver.page_source)
            except Exception as e:
                pipeline._log(job_name, f"    ⚠ 失败: {e}")
            time.sleep(random.uniform(2.5, 4.0))

        pipeline._log(job_name, f"✓ 抓取完成，共 {len(task_pool)} 份")
        # 记录本次批次，供"评估本批次"使用
        _scrape_sessions[job_name] = {t["id"] for t in task_pool}
        return True
    except Exception as e:
        logger.exception(e)
        pipeline._log(job_name, f"⚠ 抓取异常: {e}")
        pipeline._set_state(job_name, error=str(e))
        return False


# =============================================================================
# API：Prompt 编辑（含沙盒测试）
# =============================================================================

PROMPT_VERSIONS_DIR = os.path.join(ROOT, "prompt_versions")


def _list_prompt_versions(job_name):
    d = os.path.join(PROMPT_VERSIONS_DIR, _safe_dir(job_name))
    if not os.path.exists(d):
        return []
    files = sorted([f for f in os.listdir(d) if f.endswith(".txt")],
                   reverse=True)
    return [{"timestamp": f.replace(".txt", ""), "filename": f}
            for f in files[:30]]


def _backup_prompt(job_name, prompt_text):
    import datetime
    d = os.path.join(PROMPT_VERSIONS_DIR, _safe_dir(job_name))
    os.makedirs(d, exist_ok=True)
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(d, f"{ts}.txt")
    with open(path, "w", encoding="utf-8") as f:
        f.write(prompt_text)
    return ts


REQUIRED_PROMPT_KEYS = []  # P4-B: OUTPUT_INSTRUCTION is appended dynamically; context-only prompts are valid


# =============================================================================
# API：声明式规则引擎（Q6）
# =============================================================================

@app.route("/api/rules/list")
def api_rules_list():
    """返回岗位规则配置（若无则返回兼容生成的默认规则）。"""
    job_name = request.args.get("job_name", "")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name"), 400
    import rules as rule_engine
    rule_list = rule_engine.get_job_rules(job_name)
    return jsonify(ok=True, rules=rule_list)


@app.route("/api/rules/save", methods=["POST"])
def api_rules_save():
    """保存岗位规则配置（全量替换）。响应中含 prefilter_count 供前端提示。"""
    data = request.get_json() or {}
    job_name = data.get("job_name", "")
    rules_data = data.get("rules")
    if not job_name or rules_data is None:
        return jsonify(ok=False, message="缺少 job_name 或 rules"), 400
    import rules as rule_engine
    ok = rule_engine.save_job_rules(job_name, rules_data)
    if ok:
        job = database.get_job(job_name)
        pf_count = database.count_prefilter_rejects(job["id"]) if job else 0
        return jsonify(ok=True, message=f"已保存 {len(rules_data)} 条规则",
                       prefilter_count=pf_count)
    return jsonify(ok=False, message="保存失败"), 500


@app.route("/api/rules/clear_prefilter_rejects", methods=["POST"])
def api_rules_clear_prefilter_rejects():
    """
    清除该岗位的历史预筛拒绝记录，可选立即触发重新评估。
    输入: {job_name, reeval: bool}
    输出: {ok, message, reeval_started}
    """
    data = request.get_json() or {}
    job_name = data.get("job_name", "")
    do_reeval = bool(data.get("reeval", False))
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name"), 400
    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message="岗位不存在"), 404
    deleted = database.delete_prefilter_rejects_for_job(job["id"])
    if do_reeval:
        pipeline.reset_log(job_name)
        pipeline.reset_state(job_name)
        pipeline.run_evaluation_in_thread(job_name, force_reevaluate=False)
        return jsonify(ok=True, message=f"已清除 {deleted} 条预筛记录，评估任务已启动",
                       reeval_started=True)
    return jsonify(ok=True, message=f"已清除 {deleted} 条预筛记录",
                   reeval_started=False)


@app.route("/api/rules/parse_nl", methods=["POST"])
def api_rules_parse_nl():
    """
    用自然语言描述一条规则，由 LLM 解析为结构化规则 JSON。
    输入: {job_name, nl_text}
    输出: {ok, rule}   # rule 符合 rules.py 格式
    """
    data = request.get_json() or {}
    job_name = data.get("job_name", "")
    nl_text = data.get("nl_text", "").strip()
    if not nl_text:
        return jsonify(ok=False, message="缺少 nl_text"), 400

    system_prompt = """你是一个简历筛选规则解析器。
用户会用自然语言描述一条简历筛选规则，你需要将其解析为以下 JSON 格式：

```json
{
  "id": "唯一短ID（英文+下划线）",
  "label": "简短中文描述",
  "field": "简历字段名",
  "op": "运算符",
  "value": "比较值",
  "fail_reason": "不满足时的中文说明",
  "type": "hard",
  "enabled": true
}
```

可用字段（field）：
- age（整数）
- first_degree（字符串，如"全日制本科"）
- total_work_years（整数）
- certifications（字符串列表）
- industry_tags（字符串列表）
- major（字符串）

可用运算符（op）：
- lte/gte/lt/gt/eq/neq — 数值或字符串比较
- in/not_in — 值在/不在列表中
- contains_any/contains_all — 列表字段包含任意/全部
- regex — 正则匹配

type 固定为 "hard"（软性条件请写在 AI Prompt 里，不要生成规则）。
只输出 JSON，不要任何解释文字。"""

    user_prompt = f"请将以下规则解析为 JSON：\n{nl_text}"
    try:
        client = llm_utils.initialize_client()
        if not client:
            return jsonify(ok=False, message="LLM 不可用"), 500
        completion = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
            max_tokens=400,
            temperature=0.1,
        )
        raw = completion.choices[0].message.content.strip()
        # 提取 JSON
        if "```" in raw:
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        rule = json.loads(raw.strip())
        return jsonify(ok=True, rule=rule)
    except Exception as e:
        return jsonify(ok=False, message=f"解析失败: {e}"), 500


@app.route("/api/rules/test", methods=["POST"])
def api_rules_test():
    """
    测试规则配置对某个候选人的效果（沙箱测试）。
    输入: {job_name, rules, resume_id}
    输出: {ok, result}
    """
    data = request.get_json() or {}
    resume_id = data.get("resume_id", "")
    rules_data = data.get("rules", [])
    if not resume_id:
        return jsonify(ok=False, message="缺少 resume_id"), 400
    cand = database.get_candidate_by_resume_id(resume_id)
    if not cand:
        return jsonify(ok=False, message="候选人不存在"), 404
    try:
        structured = json.loads(cand.get("structured_json") or "{}")
    except Exception:
        structured = {}
    import rules as rule_engine
    result = rule_engine.apply_rules(structured, rules_data)
    return jsonify(ok=True, result=result)


@app.route("/api/rules/preview", methods=["POST"])
def api_rules_preview():
    """
    批量测试：将当前规则应用于该岗位所有已有候选人，返回通过/拒绝统计。
    输入: {job_name, rules}
    输出: {ok, total, pass_count, fail_count, fails:[{name,resume_id,reasons}]}
    """
    data = request.get_json() or {}
    job_name = data.get("job_name", "")
    rules_data = data.get("rules", [])
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name"), 400

    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message="岗位不存在"), 404

    import rules as rule_engine

    with database.get_conn() as conn:
        rows = conn.execute("""
            SELECT DISTINCT c.name, c.resume_id, c.structured_json
            FROM candidates c
            WHERE c.structured_json IS NOT NULL
              AND (
                EXISTS (
                    SELECT 1 FROM evaluations e
                    WHERE e.candidate_id = c.id AND e.job_id = ?
                )
                OR EXISTS (
                    SELECT 1 FROM prefilter_rejects pr
                    WHERE pr.resume_id = c.resume_id AND pr.job_id = ?
                )
              )
        """, (job["id"], job["id"])).fetchall()
        rows = [dict(r) for r in rows]

    if not rows:
        return jsonify(ok=True, total=0, pass_count=0, fail_count=0, fails=[],
                       message="该岗位暂无已评估候选人可供测试")

    passes, fails = 0, []
    for r in rows:
        try:
            structured = json.loads(r["structured_json"])
        except Exception:
            structured = {}
        result = rule_engine.apply_rules(structured, rules_data)
        if result["passed"]:
            passes += 1
        else:
            fails.append({
                "name": r["name"] or "未知",
                "resume_id": r["resume_id"],
                "reasons": result["hard_fail_reasons"],
            })

    return jsonify(ok=True, total=len(rows), pass_count=passes,
                   fail_count=len(fails), fails=fails)


@app.route("/api/rules/generate_from_jd", methods=["POST"])
def api_rules_generate_from_jd():
    """
    从 JD/招聘需求文本生成预览规则 + Prompt 补充建议（不自动保存）。
    输入: {job_name, jd_text}
    输出: {ok, rules, prompt_additions}
    """
    data = request.get_json() or {}
    jd_text = data.get("jd_text", "").strip()
    if not jd_text:
        return jsonify(ok=False, message="缺少 jd_text"), 400

    system_prompt = """你是一个招聘需求解析器。将用户提供的 JD 或岗位需求拆分为两类：

【硬性规则】：客观可量化、代码可精确判断的条件，输出为 rules JSON 数组。
  只用以下字段：age / first_degree / total_work_years / certifications / industry_tags / major
  只用以下运算符：lte（≤）/ gte（≥）/ in（必须是列表中之一）/ contains_any（包含任意一个）/ not_in
  学历值只能是：全日制本科 / 全日制硕士 / 全日制博士
  原则：宁可少生成，不要把"优先"或"加分项"误判为硬性规则。

【AI评估补充】：需要语境理解、无法精确量化的软性条件，输出为 prompt_additions 文字段落，
  直接可以追加到 AI 评估 Prompt 末尾，说明额外关注点。

输出纯 JSON，不加任何解释：
{
  "rules": [
    {
      "id": "唯一英文ID",
      "label": "中文描述",
      "field": "字段名",
      "op": "运算符",
      "value": "比较值（数字不加引号）",
      "fail_reason": "不满足时的说明",
      "type": "hard",
      "enabled": true
    }
  ],
  "prompt_additions": "段落文字，说明AI评估时需额外关注的软性维度"
}"""

    try:
        client = llm_utils.initialize_client()
        if not client:
            return jsonify(ok=False, message="LLM 不可用"), 500
        completion = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"请解析以下招聘需求：\n\n{jd_text}"},
            ],
            max_tokens=1000,
            temperature=0.1,
            response_format={"type": "json_object"},
        )
        raw = completion.choices[0].message.content.strip()
        parsed = json.loads(raw)
        raw_rules = parsed.get("rules", [])
        prompt_additions = parsed.get("prompt_additions", "")

        if not isinstance(raw_rules, list):
            raw_rules = []

        ALLOWED_FIELDS = {"age", "first_degree", "total_work_years",
                          "certifications", "industry_tags", "major"}
        ALLOWED_OPS = {"lte", "gte", "in", "not_in", "contains_any"}

        valid_rules, skipped = [], []
        for r in raw_rules:
            if not isinstance(r, dict):
                skipped.append("（非对象）")
                continue
            missing = [k for k in ("field", "op", "value") if k not in r or r[k] is None]
            if missing:
                skipped.append(r.get("label") or r.get("id") or "未知规则")
                continue
            if r["field"] not in ALLOWED_FIELDS:
                skipped.append(f'{r.get("label","?")}（字段 {r["field"]} 不支持）')
                continue
            if r["op"] not in ALLOWED_OPS:
                skipped.append(f'{r.get("label","?")}（运算符 {r["op"]} 不支持）')
                continue
            r.setdefault("id", f"jd_{r['field']}_{r['op']}")
            r.setdefault("label", r["id"])
            r.setdefault("fail_reason", f'{r["label"]} 不满足要求')
            r["type"] = "hard"
            r.setdefault("enabled", True)
            valid_rules.append(r)

        return jsonify(ok=True, rules=valid_rules, prompt_additions=prompt_additions,
                       skipped=skipped)
    except Exception as e:
        return jsonify(ok=False, message=f"生成失败: {e}"), 500


@app.route("/api/prompt/save", methods=["POST"])
def api_prompt_save():
    data = request.get_json() or {}
    job_name = data.get("job_name")
    prompt_text = data.get("prompt_text") or ""
    force = bool(data.get("force", False))

    if not job_name:
        return jsonify(ok=False, message="缺少 job_name")

    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message=f"岗位 {job_name} 不存在")

    missing = [k for k in REQUIRED_PROMPT_KEYS if k not in prompt_text]
    if missing and not force:
        return jsonify(ok=False, needs_confirm=True,
                       message=f"prompt 中缺少关键字段：{', '.join(missing)}。"
                               f"这可能导致 AI 输出无法解析。"
                               f"如确认要保存，请勾选「强制保存」。")

    _backup_prompt(job_name, job["prompt_text"] or "")

    config = json.loads(job["config_json"])
    job_dir = os.path.join(ROOT, "job_configs", _safe_dir(job_name))
    os.makedirs(job_dir, exist_ok=True)
    with open(os.path.join(job_dir, "prompt.txt"), "w", encoding="utf-8") as f:
        f.write(prompt_text)

    database.upsert_job(job_name, config, prompt_text)

    return jsonify(ok=True, message="已保存，旧版本已自动备份")


@app.route("/api/prompt/restore", methods=["POST"])
def api_prompt_restore():
    data = request.get_json() or {}
    job_name = data.get("job_name")
    timestamp = data.get("timestamp")

    if not job_name:
        return jsonify(ok=False, message="缺少 job_name")

    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message=f"岗位 {job_name} 不存在")

    if timestamp:
        path = os.path.join(PROMPT_VERSIONS_DIR, _safe_dir(job_name),
                            f"{timestamp}.txt")
        if not os.path.exists(path):
            return jsonify(ok=False, message="备份文件不存在")
        with open(path, "r", encoding="utf-8") as f:
            prompt_text = f.read()
    else:
        prompt_text = None
        for job_def in init_jobs.JOBS:
            if job_def["name"] == job_name:
                prompt_text = init_jobs.render_prompt(job_def)
                break
        if prompt_text is None:
            return jsonify(ok=False,
                           message="找不到该岗位的出厂默认（可能是手动新增的）")

    _backup_prompt(job_name, job["prompt_text"] or "")

    config = json.loads(job["config_json"])
    job_dir = os.path.join(ROOT, "job_configs", _safe_dir(job_name))
    os.makedirs(job_dir, exist_ok=True)
    with open(os.path.join(job_dir, "prompt.txt"), "w", encoding="utf-8") as f:
        f.write(prompt_text)
    database.upsert_job(job_name, config, prompt_text)

    return jsonify(ok=True, prompt_text=prompt_text,
                   message="已恢复，当前 prompt 已备份")


@app.route("/api/prompt/test", methods=["POST"])
def api_prompt_test():
    """沙盒测试：用草稿 Prompt 在已有候选人中抽 5 份评估，不写库"""
    data = request.get_json() or {}
    job_name = data.get("job_name")
    prompt_draft = data.get("prompt_text", "")
    sample_size = int(data.get("sample_size", 5))
    if not job_name or not prompt_draft.strip():
        return jsonify(ok=False, message="缺少 job_name 或 prompt_text")

    try:
        result = pipeline.sandbox_test(job_name, prompt_draft, sample_size=sample_size)
        return jsonify(result)
    except Exception as e:
        logger.exception(e)
        return jsonify(ok=False, message=str(e))


@app.route("/api/prompt/preview/<path:job_name>")
def api_prompt_preview(job_name):
    """返回最终注入 AI 的完整 Prompt 预览（含规则段、细则段）。"""
    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message="岗位不存在"), 404
    clean = judger._clean_job_prompt(job.get("prompt_text") or "")
    rules = judger.build_rules_section(job_name)
    criteria = judger.build_criteria_notes_section(job_name)
    full = f"{clean}\n\n{rules}\n\n{criteria}\n\n# 输入数据说明\n（候选人 JSON 将在此插入）\n\n{judger.OUTPUT_INSTRUCTION}"
    return jsonify(ok=True, preview=full.strip())


@app.route("/api/jobs/update_eval_rules", methods=["POST"])
def api_jobs_update_eval_rules():
    """保存岗位 AI 评判规则（config_json.rules），供 prompt_editor 使用。"""
    data = request.get_json() or {}
    job_name = data.get("job_name", "")
    rules = data.get("rules")
    if not job_name or rules is None:
        return jsonify(ok=False, message="缺少 job_name 或 rules"), 400
    ok = database.update_job_eval_rules(job_name, rules)
    if ok:
        return jsonify(ok=True, message=f"已保存 {len(rules)} 条评判规则")
    return jsonify(ok=False, message="岗位不存在"), 404


@app.route("/api/rules/eval_list")
def api_rules_eval_list():
    """返回岗位 AI 评判规则（config_json.rules）列表。"""
    job_name = request.args.get("job_name", "")
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name"), 400
    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message="岗位不存在"), 404
    try:
        config = json.loads(job.get("config_json") or "{}")
    except Exception:
        config = {}
    return jsonify(ok=True, rules=config.get("rules", []))


# =============================================================================
# =============================================================================
# API：清理助手（语义分组）
# =============================================================================

@app.route("/api/cleanup/groups/<path:job_name>")
def api_cleanup_groups(job_name):
    """对该岗位 verdict=排除 候选人按淘汰原因聚类分组"""
    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message="岗位不存在"), 404

    # 既要包含 evaluations 中 verdict=排除 的，也要包含 prefilter_rejects 中的
    groups = {}

    # prefilter rejects
    with database.get_conn() as conn:
        rows = conn.execute("""
            SELECT id, resume_id, fail_reason, name_hint
            FROM prefilter_rejects WHERE job_id=?
        """, (job["id"],)).fetchall()
        for r in rows:
            r = dict(r)
            key = _normalize_reason(r["fail_reason"])
            groups.setdefault(key, {
                "label": _short_label(key),
                "kind": "prefilter",
                "items": [],
            })["items"].append({
                "kind": "prefilter",
                "id": r["id"],
                "resume_id": r["resume_id"],
                "name": r["name_hint"] or "",
                "reason": r["fail_reason"],
            })

        # 评估为排除的（旧数据）
        rows = conn.execute("""
            SELECT e.id AS evaluation_id, e.verdict, e.cons_json, e.mismatches_json,
                   c.id AS candidate_id, c.resume_id, c.name
            FROM evaluations e JOIN candidates c ON e.candidate_id=c.id
            WHERE e.job_id=? AND e.verdict='排除'
        """, (job["id"],)).fetchall()
        for r in rows:
            r = dict(r)
            reason_text = ""
            try:
                mm = json.loads(r["mismatches_json"]) if r.get("mismatches_json") else []
                if mm:
                    parts = []
                    for it in mm:
                        if isinstance(it, dict):
                            parts.append(it.get("条件") or it.get("原因") or "")
                    reason_text = "; ".join([p for p in parts if p])[:80]
            except Exception:
                pass
            if not reason_text and r.get("cons_json"):
                try:
                    cons = json.loads(r["cons_json"])
                    reason_text = "; ".join(str(c) for c in cons)[:80]
                except Exception:
                    pass

            key = _normalize_reason(reason_text)
            groups.setdefault(key, {
                "label": _short_label(key),
                "kind": "evaluation",
                "items": [],
            })["items"].append({
                "kind": "evaluation",
                "evaluation_id": r["evaluation_id"],
                "candidate_id": r["candidate_id"],
                "resume_id": r["resume_id"],
                "name": r["name"] or "",
                "reason": reason_text,
            })

    out = []
    for key, g in groups.items():
        out.append({
            "key": key,
            "label": g["label"],
            "count": len(g["items"]),
            "items": g["items"][:50],
        })
    out.sort(key=lambda x: -x["count"])
    return jsonify(ok=True, groups=out)


def _normalize_reason(reason):
    """提取原因的关键词作为分组 key"""
    if not reason:
        return "其他"
    r = reason.strip()
    keywords = [
        ("学历", "学历不达标"),
        ("工作年限", "工龄不足"),
        ("年限", "工龄不足"),
        ("行业", "行业不符"),
        ("英语", "英语不达标"),
        ("应届", "应届生"),
        ("跳槽", "跳槽过频"),
        ("管理", "管理经验缺失"),
    ]
    for kw, label in keywords:
        if kw in r:
            return label
    # 取前 10 个字作为组名
    return r[:10] or "其他"


def _short_label(key):
    return key


@app.route("/api/cleanup/delete_group", methods=["POST"])
def api_cleanup_delete_group():
    """删除一组（kind=prefilter / evaluation）"""
    data = request.get_json() or {}
    job_name = data.get("job_name")
    items = data.get("items") or []
    if not job_name or not items:
        return jsonify(ok=False, message="参数缺失"), 400

    deleted = 0
    for it in items:
        try:
            if it.get("kind") == "prefilter":
                with database.get_conn() as conn:
                    conn.execute("DELETE FROM prefilter_rejects WHERE id=?",
                                 (int(it["id"]),))
                deleted += 1
            else:
                cand_id = it.get("candidate_id")
                if cand_id:
                    cand = database.get_candidate_by_id(int(cand_id))
                    database.delete_candidate(int(cand_id))
                    deleted += 1
        except Exception as e:
            logger.warning(f"删除失败 {it}: {e}")
    return jsonify(ok=True, deleted=deleted)


# =============================================================================
# API：淘汰原因标签
# =============================================================================

@app.route("/api/rejection_tags/stats")
def api_rejection_tag_stats():
    job_name = request.args.get("job_name")
    job_id = None
    if job_name:
        job = database.get_job(job_name)
        if job:
            job_id = job["id"]
    items = database.list_rejection_tag_stats(job_id=job_id, threshold=1)
    return jsonify(ok=True, items=items)


@app.route("/api/rejection_tags/options")
def api_rejection_tag_options():
    return jsonify(ok=True, options=database.get_distinct_rejection_tags(50))


# =============================================================================
# API：数据库查看器
# =============================================================================

PAGE_SIZE = 30


def _query_database_table(tab, page=1, keyword=""):
    page = max(1, int(page))
    offset = (page - 1) * PAGE_SIZE
    kw = f"%{keyword}%" if keyword else None

    with database.get_conn() as conn:
        if tab == "jobs":
            sql_count = "SELECT COUNT(*) AS c FROM jobs"
            sql_data = ("SELECT id, name, english_required, min_education, "
                        "min_years, industry_required, requires_management, "
                        "created_at, updated_at FROM jobs ORDER BY id")
            args = ()
            if kw:
                sql_count = "SELECT COUNT(*) AS c FROM jobs WHERE name LIKE ?"
                sql_data = ("SELECT id, name, english_required, min_education, "
                            "min_years, industry_required, requires_management, "
                            "created_at, updated_at FROM jobs WHERE name LIKE ? "
                            "ORDER BY id")
                args = (kw,)
            total = conn.execute(sql_count, args).fetchone()["c"]
            rows = conn.execute(sql_data + f" LIMIT {PAGE_SIZE} OFFSET {offset}",
                                args).fetchall()
        elif tab == "candidates":
            sql_count = "SELECT COUNT(*) AS c FROM candidates"
            sql_data = ("SELECT id, resume_id, name, age, first_degree, school, "
                        "major, english_level, total_years, created_at "
                        "FROM candidates ORDER BY id DESC")
            args = ()
            if kw:
                sql_count = ("SELECT COUNT(*) AS c FROM candidates "
                             "WHERE name LIKE ? OR school LIKE ? OR major LIKE ?")
                sql_data = ("SELECT id, resume_id, name, age, first_degree, school, "
                            "major, english_level, total_years, created_at "
                            "FROM candidates WHERE name LIKE ? OR school LIKE ? "
                            "OR major LIKE ? ORDER BY id DESC")
                args = (kw, kw, kw)
            total = conn.execute(sql_count, args).fetchone()["c"]
            rows = conn.execute(sql_data + f" LIMIT {PAGE_SIZE} OFFSET {offset}",
                                args).fetchall()
        elif tab == "evaluations":
            sql_count = "SELECT COUNT(*) AS c FROM evaluations"
            sql_data = ("SELECT e.id, c.name AS candidate_name, j.name AS job_name, "
                        "e.verdict, e.has_hard_fail, e.evaluated_at "
                        "FROM evaluations e "
                        "JOIN candidates c ON e.candidate_id=c.id "
                        "JOIN jobs j ON e.job_id=j.id "
                        "ORDER BY e.id DESC")
            args = ()
            if kw:
                sql_count = ("SELECT COUNT(*) AS c FROM evaluations e "
                             "JOIN candidates c ON e.candidate_id=c.id "
                             "JOIN jobs j ON e.job_id=j.id "
                             "WHERE c.name LIKE ? OR j.name LIKE ? OR e.verdict LIKE ?")
                sql_data = ("SELECT e.id, c.name AS candidate_name, j.name AS job_name, "
                            "e.verdict, e.has_hard_fail, e.evaluated_at "
                            "FROM evaluations e "
                            "JOIN candidates c ON e.candidate_id=c.id "
                            "JOIN jobs j ON e.job_id=j.id "
                            "WHERE c.name LIKE ? OR j.name LIKE ? OR e.verdict LIKE ? "
                            "ORDER BY e.id DESC")
                args = (kw, kw, kw)
            total = conn.execute(sql_count, args).fetchone()["c"]
            rows = conn.execute(sql_data + f" LIMIT {PAGE_SIZE} OFFSET {offset}",
                                args).fetchall()
        elif tab == "outcomes":
            sql_count = "SELECT COUNT(*) AS c FROM outcomes"
            sql_data = ("SELECT o.id, c.name AS candidate_name, j.name AS job_name, "
                        "o.action, o.action_at, o.note "
                        "FROM outcomes o "
                        "JOIN evaluations e ON o.evaluation_id=e.id "
                        "JOIN candidates c ON e.candidate_id=c.id "
                        "JOIN jobs j ON e.job_id=j.id "
                        "ORDER BY o.id DESC")
            args = ()
            if kw:
                sql_count = ("SELECT COUNT(*) AS c FROM outcomes o "
                             "JOIN evaluations e ON o.evaluation_id=e.id "
                             "JOIN candidates c ON e.candidate_id=c.id "
                             "JOIN jobs j ON e.job_id=j.id "
                             "WHERE c.name LIKE ? OR j.name LIKE ? OR o.action LIKE ?")
                sql_data = ("SELECT o.id, c.name AS candidate_name, j.name AS job_name, "
                            "o.action, o.action_at, o.note "
                            "FROM outcomes o "
                            "JOIN evaluations e ON o.evaluation_id=e.id "
                            "JOIN candidates c ON e.candidate_id=c.id "
                            "JOIN jobs j ON e.job_id=j.id "
                            "WHERE c.name LIKE ? OR j.name LIKE ? OR o.action LIKE ? "
                            "ORDER BY o.id DESC")
                args = (kw, kw, kw)
            total = conn.execute(sql_count, args).fetchone()["c"]
            rows = conn.execute(sql_data + f" LIMIT {PAGE_SIZE} OFFSET {offset}",
                                args).fetchall()
        else:
            return {"rows": [], "total": 0, "page": 1, "page_size": PAGE_SIZE,
                    "total_pages": 0, "keyword": keyword, "headers": []}

    rows = [dict(r) for r in rows]
    headers = list(rows[0].keys()) if rows else []
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    return {
        "rows": rows,
        "headers": headers,
        "total": total,
        "page": page,
        "page_size": PAGE_SIZE,
        "total_pages": total_pages,
        "keyword": keyword,
    }


@app.route("/api/database/delete", methods=["POST"])
def api_db_delete():
    data = request.get_json() or {}
    table = data.get("table")
    rid = data.get("id")
    if not table or not rid:
        return jsonify(ok=False, message="缺少参数")
    try:
        if table == "evaluation":
            n = database.delete_evaluation(int(rid))
        elif table == "candidate":
            cand = database.get_candidate_by_id(int(rid))
            n = database.delete_candidate(int(rid))
        else:
            return jsonify(ok=False, message=f"不允许删除 {table}")
        return jsonify(ok=True, deleted=n)
    except Exception as e:
        return jsonify(ok=False, message=str(e)), 500


# =============================================================================
# 手动导入简历（自投简历系统）
# =============================================================================

UPLOAD_TMP_DIR = os.path.join(ROOT, "data", "upload_tmp")
os.makedirs(UPLOAD_TMP_DIR, exist_ok=True)


@app.route("/api/manual_import/upload", methods=["POST"])
def api_manual_import_upload():
    """接收上传的简历文件，提取文字并用LLM解析结构化信息，返回预览。"""
    if not HAS_FILE_PARSER:
        return jsonify(ok=False, message="file_parser 模块未加载"), 500

    job_name = request.form.get("job_name", "").strip()
    if not job_name:
        return jsonify(ok=False, message="缺少 job_name"), 400

    files = request.files.getlist("files")
    if not files:
        return jsonify(ok=False, message="未收到文件"), 400

    client = llm_utils.initialize_client()
    if not client:
        return jsonify(ok=False, message="LLM 客户端不可用，请检查 .env"), 500

    previews = []
    for f in files:
        filename = f.filename or "未知文件"
        ext = os.path.splitext(filename)[1].lower()
        if ext not in (".pdf", ".docx", ".doc"):
            previews.append({"filename": filename, "error": f"不支持的格式 {ext}"})
            continue

        # 保存临时文件
        tmp_path = os.path.join(UPLOAD_TMP_DIR, f"tmp_{os.urandom(8).hex()}{ext}")
        try:
            f.save(tmp_path)

            # 计算 MD5 作为 resume_id
            with open(tmp_path, "rb") as fh:
                resume_id = hashlib.md5(fh.read()).hexdigest()

            # 提取文字
            text = file_parser.extract_text_from_file(tmp_path)
            text_preview = text[:300].replace("\n", " ") if text else ""

            # LLM 解析结构化信息
            structured = extractor.extract_with_retry(client, text)
            if not structured:
                structured = {}

            previews.append({
                "resume_id": resume_id,
                "filename": filename,
                "name": structured.get("name") or "",
                "age": structured.get("age"),
                "first_degree": structured.get("first_degree") or "",
                "school": structured.get("school") or "",
                "major": structured.get("major") or "",
                "total_years": structured.get("total_work_years"),
                "text_preview": text_preview,
                "file_temp_path": tmp_path,
                "structured_json": json.dumps(structured, ensure_ascii=False),
            })
        except Exception as e:
            logger.exception(e)
            try:
                os.remove(tmp_path)
            except Exception:
                pass
            previews.append({"filename": filename, "error": str(e)[:200]})

    return jsonify(ok=True, previews=previews)


@app.route("/api/manual_import/confirm", methods=["POST"])
def api_manual_import_confirm():
    """确认导入：移动文件、保存文字HTML、写库。"""
    data = request.get_json() or {}
    job_name = data.get("job_name", "").strip()
    items = data.get("items", [])
    _allowed_import_types = {"自投", "主动搜索"}
    import_type = data.get("import_type", "自投")
    if import_type not in _allowed_import_types:
        return jsonify(ok=False, message="无效的导入类型"), 400
    if not job_name or not items:
        return jsonify(ok=False, message="参数不完整"), 400

    job = database.get_job(job_name)
    if not job:
        return jsonify(ok=False, message="岗位不存在"), 404
    job_id = job["id"]

    _safe_upload_dir = os.path.realpath(UPLOAD_TMP_DIR) + os.sep
    _safe_ext = {".pdf", ".docx", ".doc"}

    count = 0
    errors = []
    for item in items:
        resume_id = item.get("resume_id", "")
        filename  = item.get("filename", "")
        tmp_path  = item.get("file_temp_path", "")
        structured_json_str = item.get("structured_json", "{}")

        if not resume_id or not tmp_path:
            errors.append(f"{filename}: 缺少必要参数")
            continue

        # 校验 resume_id 格式（防路径穿越）
        if not _RESUME_ID_RE.match(str(resume_id)):
            errors.append(f"{filename}: 无效的 resume_id")
            continue

        # 校验 tmp_path 必须在 UPLOAD_TMP_DIR 内（防客户端替换为任意服务器路径）
        abs_tmp = os.path.realpath(str(tmp_path))
        if not abs_tmp.startswith(_safe_upload_dir):
            errors.append(f"{filename}: 非法文件路径")
            continue

        if not os.path.exists(abs_tmp):
            errors.append(f"{filename}: 临时文件不存在（可能已过期，请重新上传）")
            continue

        # 校验扩展名（防写入非简历文件）
        ext = os.path.splitext(filename)[1].lower()
        if ext not in _safe_ext:
            ext = os.path.splitext(abs_tmp)[1].lower()
        if ext not in _safe_ext:
            ext = ".pdf"

        try:
            dest_file = os.path.join(FILES_DIR, f"{resume_id}{ext}")
            shutil.move(abs_tmp, dest_file)

            # 保存文字为 HTML（供处理台展示）
            try:
                import file_parser as _fp
                text = _fp.extract_text_from_file(dest_file)
            except Exception:
                text = ""
            # html.escape 防止提取文字中含 HTML 标签时造成 XSS
            html_content = (
                f"<!DOCTYPE html><html><head>"
                f"<meta charset='utf-8'>"
                f"<style>body{{font-family:sans-serif;font-size:13px;line-height:1.7;"
                f"padding:20px;white-space:pre-wrap}}</style></head>"
                f"<body><h3>简历原文（手动导入）</h3>"
                f"<pre>{html.escape(text)}</pre></body></html>"
            )
            text_html_path = os.path.join(FILES_DIR, f"{resume_id}_text.html")
            with open(text_html_path, "w", encoding="utf-8") as fh:
                fh.write(html_content)

            # 解析结构化信息（支持前端可能已编辑的字段）
            try:
                structured = json.loads(structured_json_str)
            except Exception:
                structured = {}
            # 前端可编辑字段覆盖
            for field, key in [("name", "name"), ("age", "age"),
                                ("first_degree", "first_degree"), ("school", "school"),
                                ("major", "major"), ("total_years", "total_work_years")]:
                if item.get(field) is not None:
                    structured[key] = item[field]

            cand_info = {
                "name": item.get("name") or structured.get("name") or filename,
                "age": item.get("age") or structured.get("age"),
                "first_degree": item.get("first_degree") or structured.get("first_degree"),
                "school": item.get("school") or structured.get("school"),
                "major": item.get("major") or structured.get("major"),
                "english_level": structured.get("english_level"),
                "total_years": item.get("total_years") or structured.get("total_work_years"),
                "raw_html_path": text_html_path,
                "structured_json": structured,
                "duplicate_of_id": None,
                "source": "manual",
            }
            candidate_id = database.upsert_candidate(resume_id, cand_info)

            # 写评估记录（verdict 由前端选择：自投 或 主动搜索）
            eval_id = database.upsert_evaluation(
                candidate_id, job_id,
                verdict=import_type,
                pros=[], cons=[],
                has_hard_fail=False,
                matches=[], mismatches=[],
                verdict_reason="",
                session_id=None,
            )

            # 自动记录为通过
            database.record_outcome(eval_id, "approved")
            count += 1

        except Exception as e:
            logger.exception(e)
            errors.append(f"{filename}: {str(e)[:200]}")

    return jsonify(ok=True, count=count, errors=errors)


# =============================================================================
# 候选人全库搜索
# =============================================================================

@app.route("/search")
def search_page():
    missing = database.count_candidates_without_fingerprint()
    jobs = database.list_jobs()
    return render_template("search.html", missing_count=missing, jobs=jobs)


def _expand_keywords(query: str, client) -> list:
    """用 AI 将查询描述扩展为 3-6 个中文关键词。"""
    try:
        resp = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "将用户的招聘需求描述扩展为3-6个中文搜索关键词，"
                        "这些关键词将用于匹配候选人能力简介。"
                        "只输出关键词，用逗号分隔，不要其他文字。"
                    ),
                },
                {"role": "user", "content": query},
            ],
            max_tokens=80,
            temperature=0.1,
        )
        raw = resp.choices[0].message.content.strip()
        parts = []
        for sep in ["，", ",", "、", " "]:
            if sep in raw:
                parts = [p.strip() for p in raw.split(sep) if p.strip()]
                break
        if not parts:
            parts = [raw.strip()] if raw.strip() else []
        return [k for k in parts if k][:6] or [query]
    except Exception:
        return [query]


def _ai_rank_candidates(query: str, candidates: list, client) -> list:
    """对前30条候选人结果做 AI 精排，返回带 match_score / match_reason 的列表。"""
    if not candidates:
        return candidates
    fingerprints_text = "\n".join(
        f"{i + 1}. [{c['name']}·{c['source_job_name']}·{c['verdict']}] "
        f"{c.get('ability_fingerprint', '')}"
        for i, c in enumerate(candidates)
    )
    try:
        resp = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "你是招聘顾问，根据招聘需求对候选人能力摘要打匹配度分。"},
                {
                    "role": "user",
                    "content": (
                        f"招聘需求：{query}\n\n"
                        f"候选人列表：\n{fingerprints_text}\n\n"
                        "请对每位候选人打匹配度分（0-100），给出一句理由（15字以内）。\n"
                        "只输出 JSON 数组格式，例如：\n"
                        '[{"index":1,"score":85,"reason":"..."},{"index":2,"score":60,"reason":"..."}]\n'
                        "不要其他文字。"
                    ),
                },
            ],
            max_tokens=900,
            temperature=0.1,
        )
        raw = resp.choices[0].message.content.strip()
        try:
            scores = json.loads(raw)
        except json.JSONDecodeError:
            start, end = raw.find("["), raw.rfind("]")
            scores = json.loads(raw[start:end + 1]) if start != -1 and end != -1 else []

        score_map = {item["index"]: item for item in scores if isinstance(item, dict)}
        for i, c in enumerate(candidates):
            info = score_map.get(i + 1, {})
            c["match_score"] = info.get("score")
            c["match_reason"] = info.get("reason", "")
        candidates.sort(key=lambda c: c.get("match_score") or 0, reverse=True)
    except Exception as e:
        logger.warning(f"AI精排失败: {e}")
    return candidates


@app.route("/api/search/candidates", methods=["POST"])
def api_search_candidates():
    data = request.get_json() or {}
    query = (data.get("query") or "").strip()
    source_jobs = data.get("source_jobs") or []
    verdict_filter = data.get("verdict_filter") or []
    use_ai_ranking = bool(data.get("use_ai_ranking", False))

    if not query:
        return jsonify(ok=False, message="请输入搜索内容"), 400

    client = llm_utils.initialize_client()
    if not client:
        return jsonify(ok=False, message="LLM 不可用"), 500

    keywords = _expand_keywords(query, client)
    candidates = database.search_candidates_by_fingerprint(
        keywords,
        source_jobs=source_jobs or None,
        verdict_filter=verdict_filter or None,
    )

    # 检查文件
    resume_ids = [c["resume_id"] for c in candidates]
    has_pdf_set, has_word_set = _check_resume_files(resume_ids)
    for c in candidates:
        c["has_pdf"] = c["resume_id"] in has_pdf_set
        c["has_word"] = c["resume_id"] in has_word_set

    if use_ai_ranking and candidates:
        candidates = _ai_rank_candidates(query, candidates[:30], client)

    return jsonify(ok=True, total=len(candidates), keywords=keywords, candidates=candidates)


# ── 批量生成指纹（后台线程）─────────────────────────────────
_fp_progress: dict = {"total": 0, "done": 0, "running": False, "errors": 0}
_fp_lock = threading.Lock()


@app.route("/api/admin/generate_fingerprints", methods=["POST"])
def api_generate_fingerprints():
    if not _is_local_request():
        return jsonify(ok=False, message="仅限本机调用"), 403
    with _fp_lock:
        if _fp_progress["running"]:
            return jsonify(ok=False, message="已有生成任务在运行中"), 409
        _fp_progress.update({"total": 0, "done": 0, "running": True, "errors": 0})

    def _run():
        client = llm_utils.initialize_client()
        if not client:
            with _fp_lock:
                _fp_progress["running"] = False
            return
        candidates = database.get_candidates_without_fingerprint()
        with _fp_lock:
            _fp_progress["total"] = len(candidates)
        for cand in candidates:
            try:
                structured_raw = cand.get("structured_json")
                if not structured_raw:
                    continue
                structured = (json.loads(structured_raw)
                              if isinstance(structured_raw, str) else structured_raw)
                fp = extractor.generate_fingerprint(structured, client)
                if fp:
                    database.update_ability_fingerprint(cand["resume_id"], fp)
                    logger.info(f"指纹生成：{cand.get('name', '')} ({cand['resume_id']})")
            except Exception as e:
                logger.warning(f"指纹生成失败 ({cand.get('resume_id', '')}): {e}")
                with _fp_lock:
                    _fp_progress["errors"] += 1
            finally:
                with _fp_lock:
                    _fp_progress["done"] += 1
        with _fp_lock:
            _fp_progress["running"] = False

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return jsonify(ok=True, message="后台生成任务已启动")


@app.route("/api/admin/fingerprint_progress")
def api_fingerprint_progress():
    with _fp_lock:
        return jsonify(**_fp_progress)


# =============================================================================
# 启动
# =============================================================================

def init_app():
    database.init_db()
    database.run_startup_data_migrations()
    if not database.list_jobs():
        init_jobs.init_all_jobs()
    _ensure_queue_worker()


# =============================================================================
# 跟进台
# =============================================================================

@app.route("/tracker")
def tracker_page():
    week_start = database.get_app_setting('tracker_week_start')
    sys_candidates = database.list_tracker_system_candidates(week_start)
    manual_candidates = database.list_tracker_manual()

    # 合并所有候选人
    all_candidates = list(sys_candidates)
    for m in manual_candidates:
        all_candidates.append({**m, 'entry_source': m.get('entry_source', 'manual')})

    # 按阶段排序（按 HR_STAGE_ORDER 索引）
    order_map = {s: i for i, s in enumerate(database.HR_STAGE_ORDER)}
    all_candidates.sort(key=lambda x: order_map.get(x.get('stage', ''), 99))

    # 同名多岗位标注
    from collections import defaultdict
    name_jobs = defaultdict(list)
    for c in all_candidates:
        name_jobs[c.get('name', '')].append(c.get('job_name', ''))

    return render_template('tracker.html',
        all_candidates=all_candidates,
        name_jobs=dict(name_jobs),
        week_start=week_start or '',
        stage_order=database.HR_STAGE_ORDER,
        tracker_active=list(database.TRACKER_ACTIVE_STAGES),
        tracker_reject=list(database.TRACKER_REJECT_STAGES),
    )


@app.route("/api/tracker/start_week", methods=["POST"])
def api_tracker_start_week():
    """开始本周：设置 tracker_week_start 为本周周一 00:00"""
    # 幂等：若本周已设置，返回现有值
    now = datetime.now()
    this_monday = (now - timedelta(days=now.weekday())).replace(
        hour=0, minute=0, second=0, microsecond=0)
    monday_iso = this_monday.isoformat()

    current = database.get_app_setting('tracker_week_start')
    if current and current >= monday_iso:
        return jsonify(ok=True, already=True, week_start=current)

    database.set_app_setting('tracker_week_start', monday_iso)
    return jsonify(ok=True, already=False, week_start=monday_iso)


@app.route("/api/tracker/start_week/preview", methods=["GET"])
def api_tracker_start_week_preview():
    """预览将被清除的候选人列表"""
    week_start = database.get_app_setting('tracker_week_start')
    now = datetime.now()
    this_monday = (now - timedelta(days=now.weekday())).replace(
        hour=0, minute=0, second=0, microsecond=0).isoformat()

    sys_all = database.list_tracker_system_candidates(week_start_iso=None)
    to_clear = []
    for c in sys_all:
        if c['stage'] not in (database.TRACKER_REJECT_STAGES | {'已入职'}):
            continue
        try:
            hr = database.get_hr_stage(c['evaluation_id'])
            times = json.loads((hr.get('stage_times_json') if hr else '') or '{}')
        except Exception:
            times = {}
        ts = times.get(c['stage'], '')
        if ts and ts < this_monday:
            to_clear.append({'name': c['name'], 'job': c['job_name'], 'stage': c['stage'], 'time': ts})
    return jsonify(ok=True, to_clear=to_clear, new_week_start=this_monday)


@app.route("/api/tracker/manual", methods=["POST"])
def api_tracker_manual_add():
    data = request.get_json() or {}
    entry_id = database.add_tracker_manual(
        name=data.get('name','').strip(),
        job_name=data.get('job_name',''),
        age=data.get('age'),
        stage=data.get('stage',''),
        stage_time=data.get('stage_time',''),
        department=data.get('department',''),
        education=data.get('education',''),
        school_major=data.get('school_major',''),
        source_label=data.get('source_label',''),
        note=data.get('note',''),
        entry_source=data.get('entry_source','manual'),
    )
    return jsonify(ok=True, id=entry_id)


@app.route("/api/tracker/manual/<int:entry_id>", methods=["PUT"])
def api_tracker_manual_update(entry_id):
    data = request.get_json() or {}
    allowed = {'name','job_name','age','stage','stage_time','department',
               'education','school_major','source_label','note'}
    kwargs = {k: v for k, v in data.items() if k in allowed}
    database.update_tracker_manual(entry_id, **kwargs)
    return jsonify(ok=True)


@app.route("/api/tracker/manual/<int:entry_id>", methods=["DELETE"])
def api_tracker_manual_delete(entry_id):
    database.delete_tracker_manual(entry_id)
    return jsonify(ok=True)


@app.route("/api/tracker/dept", methods=["POST"])
def api_tracker_dept():
    """更新系统候选人的部门字段（跟进台内联编辑）"""
    data = request.get_json() or {}
    eval_id    = data.get('evaluation_id')
    department = data.get('department', '')
    dept_from_yuyue = data.get('dept_from_yuyue', '')  # 已约面时记录的部门，用于冲突检测

    if not eval_id:
        return jsonify(ok=False, message="缺少 evaluation_id"), 400

    existing = database.get_hr_stage(int(eval_id))
    conflict = False
    if existing:
        orig = existing.get('department') or ''
        conflict = bool(orig) and orig != department

    database.upsert_hr_stage(
        int(eval_id),
        existing['job_id'] if existing else 0,
        stages=json.loads(existing.get('stages_json') or '[]') if existing else [],
        reject_reason=existing.get('reject_reason','') if existing else '',
        note=existing.get('note','') if existing else '',
        stage_times_json=json.loads(existing.get('stage_times_json') or '{}') if existing else {},
        department=department,
    )
    return jsonify(ok=True, conflict=conflict)


@app.route("/export/tracker")
def export_tracker():
    """生成并下载跟进台 Excel"""
    week_start = database.get_app_setting('tracker_week_start')
    sys_candidates = database.list_tracker_system_candidates(week_start)
    manual_candidates = database.list_tracker_manual()

    active, reject = [], []
    for c in sys_candidates:
        row = {
            'stage':       c.get('stage',''),
            'name':        c.get('name',''),
            'job_name':    c.get('job_name',''),
            'age':         c.get('age',''),
            'stage_time':  c.get('stage_time',''),
            'department':  c.get('department',''),
            'education':   c.get('first_degree',''),
            'source_label': '',
            'school_major': f"{c.get('school','') or ''} {c.get('major','') or ''}".strip(),
            'pdf': '',
        }
        if c['stage'] in database.TRACKER_ACTIVE_STAGES:
            active.append(row)
        else:
            reject.append(row)

    for m in manual_candidates:
        row = {
            'stage':       m.get('stage',''),
            'name':        m.get('name',''),
            'job_name':    m.get('job_name',''),
            'age':         m.get('age',''),
            'stage_time':  m.get('stage_time',''),
            'department':  m.get('department',''),
            'education':   m.get('education',''),
            'source_label': m.get('source_label',''),
            'school_major': m.get('school_major',''),
            'pdf': '',
        }
        if m.get('stage','') in database.TRACKER_REJECT_STAGES or m.get('stage','') in {'不合适','个人放弃'}:
            reject.append(row)
        else:
            active.append(row)

    order_map = {s: i for i, s in enumerate(database.HR_STAGE_ORDER)}
    active.sort(key=lambda x: order_map.get(x.get('stage',''), 99))
    reject.sort(key=lambda x: order_map.get(x.get('stage',''), 99))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "面试跟进台"

    # 生成时间范围
    now_dt = datetime.now()
    if week_start:
        try:
            ws_dt = datetime.fromisoformat(week_start)
            we_dt = ws_dt + timedelta(days=6)
            date_range = f"{ws_dt.strftime('%Y-%m-%d')} 至 {we_dt.strftime('%Y-%m-%d')}"
        except Exception:
            date_range = now_dt.strftime('%Y-%m-%d')
    else:
        date_range = now_dt.strftime('%Y-%m-%d')

    export_time = now_dt.strftime('%Y-%m-%d %H:%M')

    # 第1行：元信息
    ws.append([f"导出时间：{export_time}　　数据范围：{date_range}"])
    ws.merge_cells('A1:J1')
    ws['A1'].font = Font(color='888888', italic=True, size=10)

    # 表头
    headers = ['面试结果','姓名','应聘岗位','年龄','面试/入职时间','需求部门','学历','简历来源','毕业院校/专业','简历PDF']
    ws.append(headers)
    header_row = ws[2]
    for cell in header_row:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='3D5AFE')
        cell.alignment = Alignment(horizontal='center')

    # 列宽
    col_widths = [10,10,14,6,16,12,10,10,24,10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    def write_rows(rows, fill_color=None):
        for r in rows:
            ws.append([
                r['stage'], r['name'], r['job_name'], r['age'],
                r['stage_time'], r['department'], r['education'],
                r['source_label'], r['school_major'], r['pdf'],
            ])
            if fill_color:
                for cell in ws[ws.max_row]:
                    cell.fill = PatternFill('solid', fgColor=fill_color)

    # 活跃区域
    if active:
        ws.append(['── 活跃候选人 ──'])
        ws.merge_cells(f'A{ws.max_row}:J{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(bold=True, color='1565C0')
        ws[f'A{ws.max_row}'].fill = PatternFill('solid', fgColor='E3F2FD')

    write_rows([r for r in active if r['stage'] != '已入职'])
    # 已入职单独绿色
    write_rows([r for r in active if r['stage'] == '已入职'], fill_color='E8F5E9')

    # 不通过区域
    if reject:
        ws.append(['── 不通过 ──'])
        ws.merge_cells(f'A{ws.max_row}:J{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(bold=True, color='B71C1C')
        ws[f'A{ws.max_row}'].fill = PatternFill('solid', fgColor='FFEBEE')
        write_rows(reject, fill_color='F5F5F5')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # 文件名
    if week_start:
        try:
            ws_dt = datetime.fromisoformat(week_start)
            we_dt = ws_dt + timedelta(days=6)
            fname = f"面试跟进表_{ws_dt.strftime('%Y%m%d')}至{we_dt.strftime('%Y%m%d')}_导出{now_dt.strftime('%m%d-%H%M')}.xlsx"
        except Exception:
            fname = f"面试跟进表_{now_dt.strftime('%Y%m%d-%H%M')}.xlsx"
    else:
        fname = f"面试跟进表_{now_dt.strftime('%Y%m%d-%H%M')}.xlsx"

    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname,
    )


if __name__ == "__main__":
    init_app()
    print("=" * 60)
    print("  华阳精机简历筛选系统 v2")
    print("  本机访问：http://127.0.0.1:5000")
    lan = get_lan_ip()
    if lan:
        print(f"  局域网访问：http://{lan}:5000")
        print(f"  （把这个地址告诉同事，他们浏览器打开即可）")
    print("=" * 60)
    app.run(host="0.0.0.0", port=5000, debug=False, use_reloader=False)
