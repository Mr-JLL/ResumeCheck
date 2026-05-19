# -*- coding: utf-8 -*-
import openpyxl
import os
import sys
sys.stdout.reconfigure(encoding='utf-8')

f = next(ff for ff in os.listdir('.') if ff.endswith('.xlsx') and '2026' in ff and '招聘' in ff)
print(f"File: {f}")
wb = openpyxl.load_workbook(f, data_only=True, read_only=True)

# Print each sheet one at a time with separator
for sn in wb.sheetnames:
    ws = wb[sn]
    print(f"\n{'='*60}")
    print(f"SHEET: {sn}")
    print(f"Dimensions: rows={ws.max_row}, cols={ws.max_column}")
    print(f"{'='*60}")
    row_num = 0
    for row in ws.iter_rows(values_only=True):
        row_num += 1
        if any(c is not None for c in row):
            # Only print first 30 cols to avoid too long lines
            print(f"R{row_num:03d}: {row[:30]}")
        if row_num >= 150:
            print("  ... (max 150 rows per sheet shown)")
            break
wb.close()
